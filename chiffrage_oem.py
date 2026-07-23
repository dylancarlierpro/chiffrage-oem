"""
Procédure Chiffrage OEM — Anneaux / Segmentation  v3.0
=======================================================
Application PyQt6 de calcul de prix unitaires OEM pour les anneaux et
segments en PTFE/PEEK vendus à différents clients industriels.

PRINCIPE DE CALCUL
------------------
On part du prix PL21 (prix catalogue 2021) saisi par l'utilisateur,
puis on applique successivement les taux d'augmentation annuels de chaque
client (2022, 2023, 2025, 2026…), et enfin la remise négociée avec ce client.
Le résultat est arrondi au centime supérieur (math.ceil).

  prix_final = ceil( PL21 × ∏(1 + taux_année) × (1 + remise) )

STRUCTURE DES DONNÉES (fichier chiffrage_data.json)
----------------------------------------------------
{
  "augmentations": {
    "PTFE": { "NomClient": { "2022": 0.06, "2023": 0.09, ... }, ... },
    "PEEK": { ... }
  },
  "remises": {
    "PTFE": { "NomClient": -0.65, ... },   # valeur négative = réduction
    "PEEK": { ... }
  },
  "remarques":  { "NomClient": "texte libre", ... },
  "historique": [ { ts, pl21, matiere, client, prix_final }, ... ],
  "audit_log":  [ { ts, action, matiere, client, details }, ... ],
  "admin_hash": "<hash salé (pbkdf2_sha256) du mot de passe admin>"
}

ARCHITECTURE DE L'INTERFACE
----------------------------
MainWindow
├── Header (bandeau plat, couleur de matière portée par les contrôles)
├── Barre matière + PL21 (sélecteur PTFE/PEEK, saisie, Calculer, Réinitialiser, filtre)
└── QSplitter
    ├── Sidebar gauche : liste de ClientListItem (nom + prix final)
    └── ClientDetailPanel : cascade de calcul (WaterfallChart), remise, PDF, copier

DÉPENDANCES
-----------
  pip install PyQt6 openpyxl
  openpyxl est optionnel (export Excel désactivé si absent).

Au premier lancement (aucun chiffrage_data.json existant), un mot de passe
admin est généré aléatoirement et affiché une seule fois à l'écran — voir
get_admin_hash().

Créé par Dylan Carlier, imaginé par Nicolas Richet et Robin Demeure.
"""

import sys
import json
import math
import hashlib
import hmac
import secrets
import copy
import csv
import datetime
import os
import time
import unicodedata
import html as _html
from pathlib import Path
import re as _re

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QFrame,
    QGroupBox, QMessageBox, QDialog, QDialogButtonBox, QFormLayout,
    QDoubleSpinBox, QInputDialog, QScrollArea,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QSplitter, QGraphicsDropShadowEffect, QGraphicsOpacityEffect,
    QStatusBar, QToolTip,
)
from PyQt6.QtCore import (
    Qt, QTimer, QRectF, QPointF, QPoint, QSize, QEvent, QObject,
    QVariantAnimation, QPropertyAnimation, QEasingCurve,
)
from PyQt6.QtGui import (
    QFont, QColor, QPalette, QPixmap, QAction, QPainter, QPen, QBrush,
    QIcon, QPainterPath, QShortcut, QKeySequence,
)
# QTextDocument/QPrinter (export PDF) et openpyxl (export Excel) ne sont
# utilisés que dans les fonctions d'export, jamais au démarrage — importés
# localement dans ces fonctions plutôt qu'ici pour ne pas retarder le premier
# affichage de l'application avec des modules que la majorité des sessions
# n'utilisent jamais (le sous-système d'impression Qt et openpyxl pèsent
# chacun une fraction de seconde au chargement).
import importlib.util
HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


# ─────────────────────────────────────────────────────────────────────────────
#  DONNÉES — chargement, sauvegarde et valeurs par défaut
# ─────────────────────────────────────────────────────────────────────────────

def _app_dir() -> Path:
    """
    Dossier de référence pour les fichiers de données/logo.
    Une fois packagé en .exe (PyInstaller --onefile), sys.frozen vaut True et
    __file__ pointe vers le dossier temporaire d'extraction (vidé à chaque
    lancement) : on utilise alors le dossier de l'exécutable pour que
    chiffrage_data.json survive entre deux lancements.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _data_dir() -> Path:
    """
    Dossier de stockage de chiffrage_data.json.
    Une fois installée (Inno Setup installe par défaut dans Program Files),
    l'application tourne avec des droits utilisateur standard qui n'ont PAS
    le droit d'écrire dans ce dossier : une tentative de sauvegarde y lève
    une PermissionError, qui plantait l'appli au premier clic sur « Calculer ».
    On stocke donc les données dans %APPDATA% (toujours accessible en écriture),
    et on reste à côté du script en mode développement pour ne rien changer
    au confort de dev.
    """
    if getattr(sys, "frozen", False):
        base = Path(os.environ.get("APPDATA", str(Path.home()))) / "CPI Howden" / "Chiffrage OEM"
        base.mkdir(parents=True, exist_ok=True)
        return base
    return Path(__file__).parent


# Le fichier JSON est stocké dans un dossier utilisateur inscriptible (voir _data_dir).
# S'il n'existe pas au premier lancement, DEFAULT_DATA est utilisé.
DATA_FILE = _data_dir() / "chiffrage_data.json"

# Migration : si une ancienne version stockait déjà des données à côté de
# l'exécutable (comportement d'avant ce correctif), on les récupère pour ne
# pas perdre les remises/remarques déjà personnalisées par l'utilisateur.
if getattr(sys, "frozen", False) and not DATA_FILE.exists():
    _old_data_file = _app_dir() / "chiffrage_data.json"
    if _old_data_file.exists():
        try:
            import shutil
            shutil.copy(_old_data_file, DATA_FILE)
        except Exception:
            pass

# Données initiales livrées avec l'application.
# Pour ajouter un client : l'ajouter dans "augmentations" ET "remises",
# pour CHAQUE matière présente.
# Les taux sont des décimaux : 0.06 = +6 %.
# Les remises sont négatives : -0.65 = -65 %.
DEFAULT_DATA = {
    "augmentations": {
        "PTFE": {
            "Client A": {"2022": 0.05, "2023": 0.07, "2025": 0.02, "2026": 0.04},
            "Client B": {"2022": 0.00, "2023": 0.06, "2025": 0.00, "2026": 0.04},
            "Client C": {"2022": 0.06, "2023": 0.08, "2025": 0.03, "2026": 0.04},
        },
        "PEEK": {
            "Client A": {"2022": 0.05, "2023": 0.07, "2025": 0.02, "2026": 0.04},
            "Client B": {"2022": 0.00, "2023": 0.06, "2025": 0.00, "2026": 0.04},
            "Client C": {"2022": 0.06, "2023": 0.08, "2025": 0.03, "2026": 0.04},
        },
    },
    "remises": {
        "PTFE": {"Client A": -0.60, "Client B": -0.65, "Client C": -0.55},
        "PEEK": {"Client A": -0.65, "Client B": -0.70, "Client C": -0.60},
    },
    # Remarques libres affichées sous le tableau de calcul d'un client.
    # Clé = nom du client, valeur = texte.
    "remarques": {
        "Client B": "Exemple de remarque libre associée à un client.",
    },
    "historique": [],  # liste des calculs effectués (voir log_calcul)
    "audit_log":  [],  # journal des modifications admin (voir log_audit)
}


# Rempli par load_data() si le fichier de données existait mais était illisible
# (JSON corrompu) : MainWindow affiche un avertissement une fois l'UI démarrée
# (aucune QMessageBox n'est utilisable avant la création de la QApplication).
CORRUPT_DATA_BACKUP: Path | None = None

# Vrai si chiffrage_data.json existait déjà et a été chargé avec succès (par
# opposition à un tout premier lancement, où son absence est normale). Utilisé
# par get_admin_hash() pour distinguer « nouvelle installation » de « fichier
# existant dont la clé admin_hash a disparu » (suppression manuelle/tampering)
# — ce dernier cas mérite une trace dans le journal d'audit, pas l'autre.
DATA_FILE_EXISTED_AT_LOAD = False


def _reconcile_data(d: dict) -> dict:
    """
    Valide et répare la structure de 'augmentations'/'remises' après chargement.

    Un JSON peut être syntaxiquement valide mais structurellement inutilisable
    (édité à la main) : 'augmentations'/'remises' absentes ou d'un mauvais
    type, un client présent d'un côté mais pas de l'autre, un taux non
    numérique... Sans validation, ça fait planter l'application avec un
    KeyError/AttributeError/TypeError cryptique au premier client sélectionné
    (calculer_prix, ClientDetailPanel.load_client), bien après le démarrage —
    beaucoup plus difficile à diagnostiquer qu'un échec net au chargement.

    - Si 'augmentations' ou 'remises' n'est pas un dict (absent, liste,
      chaîne...) : lève ValueError, traité par l'appelant comme un fichier
      corrompu (sauvegarde .bak + repli sur DEFAULT_DATA) — impossible à
      réparer localement sans données de référence.
    - Incohérences plus fines (client sans remise, taux non numérique...) :
      réparées en place avec des valeurs neutres, pour ne perdre que le
      strict minimum plutôt que tout le fichier.
    """
    aug = d.get("augmentations")
    rem = d.get("remises")
    if not isinstance(aug, dict) or not isinstance(rem, dict):
        raise ValueError("'augmentations'/'remises' manquant ou de type invalide")
    for matiere, clients in aug.items():
        if not isinstance(clients, dict):
            raise ValueError(f"matière '{matiere}' invalide")
        if not isinstance(rem.get(matiere), dict):
            rem[matiere] = {}
        for client, taux in clients.items():
            if not isinstance(taux, dict):
                clients[client] = {}
            else:
                clients[client] = {
                    a: v for a, v in taux.items() if isinstance(v, (int, float))
                }
            if not isinstance(rem[matiere].get(client), (int, float)):
                rem[matiere][client] = -0.60   # valeur de repli neutre
    return d


def load_data() -> dict:
    """
    Charge les données depuis chiffrage_data.json.
    Si le fichier est absent, retourne une copie de DEFAULT_DATA.
    Si le fichier existe mais est corrompu (JSON invalide ou structurellement
    inutilisable, voir _reconcile_data), il est renommé en .bak (horodaté,
    pour ne jamais écraser une sauvegarde d'un incident précédent) au lieu
    d'être silencieusement abandonné — perdre les remises/remarques/historique
    négociés sans trace serait pire que garder le fichier de côté pour
    investigation. CORRUPT_DATA_BACKUP est alors renseigné pour que
    l'utilisateur en soit informé au démarrage de l'UI.
    Les clés 'historique', 'audit_log' et 'remarques' sont créées si manquantes
    (compatibilité avec d'anciens fichiers qui ne les avaient pas).
    """
    global CORRUPT_DATA_BACKUP, DATA_FILE_EXISTED_AT_LOAD
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
            d.setdefault("historique", [])
            d.setdefault("audit_log", [])
            d.setdefault("remarques", {})
            d = _reconcile_data(d)
            DATA_FILE_EXISTED_AT_LOAD = True
            return d
        except Exception:
            horodatage = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = DATA_FILE.with_suffix(f".corrompu.{horodatage}.bak")
            try:
                DATA_FILE.rename(backup)
                CORRUPT_DATA_BACKUP = backup
            except OSError:
                pass
    return copy.deepcopy(DEFAULT_DATA)


def save_data(data: dict):
    """
    Sauvegarde le dictionnaire DATA dans chiffrage_data.json (encodage UTF-8, indenté).

    Écriture atomique : on écrit d'abord dans un fichier .tmp puis on le
    remplace via os.replace() (opération atomique sur Windows comme sur
    Unix). Sans ça, un plantage, une coupure de courant ou un disque plein
    en plein milieu de l'écriture directe du fichier final laisserait un
    JSON tronqué — que load_data() prendrait ensuite pour un fichier corrompu,
    avec perte de toutes les remises/remarques/historique déjà personnalisés.

    N'importe quelle erreur d'écriture (droits insuffisants, disque plein,
    dossier supprimé…) est signalée à l'utilisateur au lieu de faire planter
    l'application : PyQt6 abandonne (abort) le processus si une exception
    traverse un slot Qt.
    """
    tmp_path = DATA_FILE.with_suffix(DATA_FILE.suffix + ".tmp")
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, DATA_FILE)
    except OSError as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        try:
            QMessageBox.warning(
                None, "Sauvegarde impossible",
                f"Les données n'ont pas pu être enregistrées :\n{e}"
            )
        except Exception:
            pass


# Variable globale unique contenant toutes les données en mémoire.
# Toutes les classes y accèdent directement — pas de passage par paramètre.
DATA = load_data()


def log_calcul(pl21: float, matiere: str, client: str, prix_final: int, persist: bool = True):
    """
    Enregistre un calcul dans l'historique (max 500 entrées, les plus anciennes supprimées).
    Appelé automatiquement à chaque clic sur « Calculer ».

    persist=False n'ajoute l'entrée qu'en mémoire, sans réécrire le fichier
    sur disque : utile quand l'appelant enregistre plusieurs clients d'affilée
    (ex. « Calculer » sur toute la liste) et préfère sauvegarder une seule
    fois à la fin plutôt qu'une fois par client.
    """
    DATA.setdefault("historique", []).append({
        "ts":         datetime.datetime.now().isoformat(timespec="seconds"),
        "pl21":       round(pl21, 2),
        "matiere":    matiere,
        "client":     client,
        "prix_final": prix_final,
    })
    # Limite à 500 entrées pour ne pas faire grossir le fichier indéfiniment
    if len(DATA["historique"]) > 500:
        DATA["historique"] = DATA["historique"][-500:]
    if persist:
        save_data(DATA)


def log_audit(action: str, matiere: str = "", client: str = "", details: str = ""):
    """
    Enregistre une action administrative dans le journal d'audit (max 1 000 entrées).
    Exemples d'actions : 'add_client', 'edit_client', 'delete_client',
                         'edit_remise', 'add_matiere', 'change_password'.
    """
    DATA.setdefault("audit_log", []).append({
        "ts":      datetime.datetime.now().isoformat(timespec="seconds"),
        "action":  action,
        "matiere": matiere,
        "client":  client,
        "details": details,
    })
    if len(DATA["audit_log"]) > 1000:
        DATA["audit_log"] = DATA["audit_log"][-1000:]
    save_data(DATA)


# ─────────────────────────────────────────────────────────────────────────────
#  EXPRESSION FLEXIBLE — interprétation de la saisie PL21
# ─────────────────────────────────────────────────────────────────────────────

def evaluer_expression(texte: str):
    """
    Interprète la saisie PL21 de l'utilisateur et retourne (valeur_float, label_str).

    Formats acceptés :
      • Nombre simple     : "620"      → (620.0, "620 = 620.00")
      • Opération simple  : "600 + 50" → (650.0, "600 + 50 = 650.00")
      • Base ± pourcentage: "620 + 3%" → (638.6, "620.0 + 3.0% = 638.60")
                            "620 - 5%" → (589.0, ...)

    Lève ValueError si le format n'est pas reconnu ou si le résultat est ≤ 0.
    Le label est affiché en prévisualisation sous le champ de saisie.
    """
    # Normalisation : virgule → point, espaces supprimés
    t = texte.strip().replace(",", ".").replace(" ", "")
    if not t:
        raise ValueError("vide")
    # Garde-fou : une saisie démesurément longue (nombre à des centaines de
    # chiffres, chaîne d'opérateurs répétés) n'a aucun sens pour un prix et
    # peut ralentir inutilement l'évaluation qui suit.
    if len(t) > 32:
        raise ValueError(f"Saisie trop longue : {texte}")

    # Cas "base ± x%" — ex: "620+3%" ou "500-10%"
    m = _re.match(r'^([0-9]+(?:\.[0-9]*)?)([+-])([0-9]+(?:\.[0-9]*)?)%$', t)
    if m:
        base = float(m.group(1))
        op   = m.group(2)
        pct  = float(m.group(3))
        result = base * (1 + pct / 100) if op == '+' else base * (1 - pct / 100)
        # Une remise > 100% (ex: "10-500%") donnerait un PL21 nul ou négatif,
        # ce qui provoquerait une division par zéro plus loin dans le calcul.
        if not math.isfinite(result) or result <= 0:
            raise ValueError(f"Résultat négatif, nul ou trop grand : {texte}")
        return result, f"{base} {op} {pct}% = {result:.2f}"

    # Cas expression arithmétique pure (chiffres + opérateurs de base).
    # '**' (puissance) est explicitement exclu : eval() est restreint à un
    # environnement sans builtins (pas d'injection de code possible), mais
    # une expression comme "9**999999999" resterait un calcul d'entier Python
    # à précision arbitraire portant sur des centaines de millions de chiffres,
    # qui gèlerait l'UI (mono-thread) pendant un temps déraisonnable.
    if '**' not in t and _re.match(r'^[0-9+\-*/.(). ]+$', t):
        try:
            result = float(eval(t, {"__builtins__": {}}))
            if not math.isfinite(result) or result <= 0:
                raise ValueError("négatif, nul ou trop grand")
            return result, f"{texte.strip()} = {result:.2f}"
        except Exception:
            raise ValueError(f"Expression invalide : {texte}")

    raise ValueError(f"Format non reconnu : {texte}")


# ─────────────────────────────────────────────────────────────────────────────
#  CALCUL DU PRIX FINAL
# ─────────────────────────────────────────────────────────────────────────────

def calculer_prix(pl21: float, matiere: str, client: str) -> dict:
    """
    Calcule le prix unitaire final pour un client donné à partir du PL21.

    Retourne un dictionnaire :
    {
      "etapes": [
          {"annee": "2022", "taux": 0.06, "prix_apres": 656.0},
          ...
      ],
      "prix_avant_remise": 756.4,   # prix après toutes les augmentations
      "remise_pct": 65,             # remise en % (entier positif, ex: 65 pour -65%)
      "prix_final": 265,            # prix arrondi au centime supérieur (math.ceil)
    }

    Les années sont triées alphabétiquement (ordre chronologique si format AAAA).
    """
    augments = DATA["augmentations"][matiere][client]
    remise   = DATA["remises"][matiere][client]           # valeur négative, ex: -0.65
    annees   = sorted(augments.keys())

    etapes = []
    prix   = pl21
    for annee in annees:
        taux = augments[annee]
        prix = prix * (1 + taux)                          # application cumulative
        etapes.append({"annee": annee, "taux": taux, "prix_apres": prix})

    return {
        "etapes":            etapes,
        "prix_avant_remise": prix,
        "remise_pct":        int(abs(remise) * 100),
        "prix_final":        math.ceil(prix * (1 + remise)),
    }


def export_html_to_pdf(parent, html: str, path: str) -> bool:
    """
    Imprime `html` en PDF vers `path` via QTextDocument + QPrinter, avec
    gestion d'erreur : QPrinter échoue souvent silencieusement côté Qt (pas
    d'exception Python) quand le fichier cible est verrouillé (ex: déjà
    ouvert dans un lecteur PDF) — on vérifie donc après coup que le fichier a
    bien été créé et n'est pas vide, en plus d'attraper les erreurs Python
    (permissions, chemin invalide…). Affiche un avertissement et retourne
    False en cas d'échec ; True si l'export a réussi.
    """
    from PyQt6.QtGui import QTextDocument
    from PyQt6.QtPrintSupport import QPrinter
    try:
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(path)
        doc = QTextDocument()
        doc.setHtml(html)
        doc.print(printer)
        ok = Path(path).is_file() and Path(path).stat().st_size > 0
    except OSError:
        ok = False
    if not ok:
        QMessageBox.warning(
            parent, "Export impossible",
            f"Le PDF n'a pas pu être écrit (peut-être déjà ouvert ailleurs) :\n{path}"
        )
    return ok


def csv_safe(value: str) -> str:
    """
    Neutralise l'injection de formule CSV/Excel (CWE-1236) : un nom de client
    commençant par =, +, -, @, tab ou CR serait interprété comme une formule
    par Excel/LibreOffice à l'ouverture du fichier exporté (ex: un nom
    "=CMD(...)" pourrait exécuter du code chez qui ouvre le fichier). On
    préfixe d'une apostrophe, qu'Excel affiche comme texte littéral et ignore
    à l'import — inoffensif pour les noms de clients légitimes.
    """
    if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def _normalize_search(s: str) -> str:
    """
    Neutralise casse, accents et espaces de bord pour une recherche tolérante
    (ex: "francois" retrouve "Atelier François", "atlas " avec espace final
    retrouve toujours "Atlas"). Utilisé par le filtre clients de la sidebar.
    """
    s = unicodedata.normalize("NFKD", s.strip().lower())
    return "".join(c for c in s if not unicodedata.combining(c))


def _fmt_remise(pct: float) -> str:
    """
    Formate un pourcentage de remise (déjà positif, ex: 65.0 pour -65 %) en
    '-65.0%' — mais '0.0%' plutôt que '-0.0%' quand la remise est nulle : avec
    f"-{pct:.1f}%", pct=0.0 produit littéralement "-0.0%", trompeur.
    """
    return f"-{pct:.1f}%" if pct else "0.0%"


# ─────────────────────────────────────────────────────────────────────────────
#  AUTHENTIFICATION ADMINISTRATEUR
# ─────────────────────────────────────────────────────────────────────────────

PBKDF2_ITERATIONS = 200_000


def _hash_password(pwd: str, salt: bytes | None = None) -> str:
    """
    Dérive un hash salé (PBKDF2-HMAC-SHA256) du mot de passe, au format
    'pbkdf2_sha256$<itérations>$<sel hex>$<hash hex>'. Un sel aléatoire par
    mot de passe empêche un brute-force par table précalculée si
    chiffrage_data.json venait à être exfiltré.
    """
    if salt is None:
        salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", pwd.encode(), salt, PBKDF2_ITERATIONS)
    return f"pbkdf2_sha256${PBKDF2_ITERATIONS}${salt.hex()}${digest.hex()}"


def _verify_password(pwd: str, stored: str) -> bool:
    """Vérifie pwd contre un hash stocké, salé (nouveau format) ou non (ancien)."""
    if stored.startswith("pbkdf2_sha256$"):
        try:
            _, iterations_s, salt_hex, hash_hex = stored.split("$")
            digest = hashlib.pbkdf2_hmac(
                "sha256", pwd.encode(), bytes.fromhex(salt_hex), int(iterations_s)
            )
            # hmac.compare_digest plutôt que == : comparaison à temps constant,
            # qui ne fuit pas la position du premier octet différent via une
            # attaque temporelle (peu probable ici, mais sans coût à corriger).
            return hmac.compare_digest(digest.hex(), hash_hex)
        except (ValueError, IndexError):
            return False
    # Ancien format (SHA-256 non salé, fichiers de données créés avant la
    # migration) : comparaison directe, pour rester compatible avec les
    # installations existantes. Migré vers le format salé au succès (voir
    # verifier_admin) plutôt que de forcer un changement immédiat.
    return hmac.compare_digest(hashlib.sha256(pwd.encode()).hexdigest(), stored)


def get_admin_hash() -> str:
    """
    Retourne le hash du mot de passe admin stocké dans DATA.
    Si aucun hash n'est défini (premier lancement), un mot de passe est
    généré aléatoirement, haché/salé et persisté immédiatement ; il est
    affiché une seule fois à l'écran pour que l'utilisateur puisse le noter
    (voir _reveal_generated_password ci-dessous). Un mot de passe par défaut
    fixe et documenté dans le code serait une porte d'entrée admin connue de
    quiconque lit la source.

    Le fichier de données n'étant pas chiffré, un utilisateur ayant accès au
    poste peut en théorie éditer chiffrage_data.json à la main pour supprimer
    la clé 'admin_hash' et forcer la génération d'un nouveau mot de passe.
    Ce n'est pas évitable sans changer fondamentalement le stockage (voir
    limite documentée en tête de fichier) — mais si le fichier existait déjà
    (DATA_FILE_EXISTED_AT_LOAD), ce n'est PAS un premier lancement normal :
    on trace l'anomalie dans le journal d'audit pour qu'elle reste visible.
    """
    stored = DATA.get("admin_hash")
    if stored is None:
        generated_pwd = secrets.token_urlsafe(9)  # ~12 caractères lisibles
        stored = _hash_password(generated_pwd)
        DATA["admin_hash"] = stored
        if DATA_FILE_EXISTED_AT_LOAD:
            log_audit(
                "admin_hash_reset",
                details="Clé admin_hash absente d'un fichier de données existant "
                        "— nouveau mot de passe généré aléatoirement."
            )
        else:
            save_data(DATA)
        _reveal_generated_password(generated_pwd)
    return stored


def _reveal_generated_password(pwd: str):
    """
    Affiche une seule fois, dans une QMessageBox, le mot de passe admin qui
    vient d'être généré aléatoirement — sinon personne ne pourrait jamais se
    connecter en admin. Appelée uniquement depuis get_admin_hash(), toujours
    déclenchée par une action utilisateur (voir verifier_admin) donc après la
    création de la QApplication.
    """
    QMessageBox.information(
        None,
        "Mot de passe admin généré",
        "Aucun mot de passe admin n'était défini pour cette installation.\n\n"
        f"Mot de passe généré : {pwd}\n\n"
        "Notez-le : il ne sera plus jamais affiché. Vous pourrez le changer "
        "depuis le menu Admin une fois connecté.",
    )


def set_admin_hash(new_hash: str):
    """
    Remplace le hash admin en mémoire et trace l'action.
    Pas de save_data() ici : log_audit() persiste déjà DATA (avec ce nouveau
    hash inclus) après avoir ajouté l'entrée d'audit — un appel supplémentaire
    écrirait le même contenu sur disque une seconde fois pour rien.
    """
    DATA["admin_hash"] = new_hash
    log_audit("change_password", details="Mot de passe admin modifié")


# Compteur d'échecs consécutifs (en mémoire, réinitialisé à chaque succès ou
# redémarrage) : ne verrouille jamais définitivement l'admin légitime, mais
# introduit un délai croissant qui rend un essai-erreur manuel dissuasif et
# laisse une trace dans le journal d'audit (voir verifier_admin).
_failed_admin_attempts = 0


def verifier_admin(parent) -> bool:
    """
    Affiche une boîte de dialogue demandant le mot de passe admin.
    Retourne True si le mot de passe est correct, False sinon.
    'parent' est la fenêtre Qt parente pour centrer la boîte de dialogue.
    """
    global _failed_admin_attempts
    pwd, ok = QInputDialog.getText(
        parent, "Accès administrateur", "Mot de passe :",
        QLineEdit.EchoMode.Password
    )
    if not ok:
        return False
    stored = get_admin_hash()
    if _verify_password(pwd, stored):
        _failed_admin_attempts = 0
        if not stored.startswith("pbkdf2_sha256$"):
            # Migration transparente vers le format salé au premier succès.
            set_admin_hash(_hash_password(pwd))
        return True
    _failed_admin_attempts += 1
    log_audit("auth_failed", details=f"Tentative {_failed_admin_attempts} échouée")
    if _failed_admin_attempts >= 3:
        # Délai croissant après le 3e échec consécutif (plafonné à 5 s) :
        # dissuade l'essai-erreur manuel sans jamais bloquer l'admin légitime.
        time.sleep(min(_failed_admin_attempts - 2, 5))
    QMessageBox.warning(parent, "Accès refusé", "Mot de passe incorrect.")
    return False


# ─────────────────────────────────────────────────────────────────────────────
#  THÈME & DESIGN — palettes clair/sombre, avatars, badges
# ─────────────────────────────────────────────────────────────────────────────

# Palette complète pour chaque thème. Toute couleur affichée dans l'interface
# principale (hors PDF, volontairement toujours clair pour l'impression) doit
# passer par theme() plutôt que d'être codée en dur, pour rester cohérente
# entre les deux modes.
#
# Palette « fiche technique matière » : fond papier chaud plutôt que gris-bleu
# SaaS, et deux couleurs de matière (voir MATERIAL_ACCENTS plus bas) au lieu
# d'un accent bleu unique — la couleur porte une information (PTFE/PEEK,
# remise, prix final) au lieu d'être purement décorative.
THEMES = {
    "light": {
        "window_bg": "#f5f2ea", "card_bg": "#ffffff", "sidebar_bg": "#efead9",
        "text": "#23241f", "text_secondary": "#6b6a5f", "text_muted": "#9a9789",
        "border": "#dbd6c7", "input_border": "#c7c0ac",
        "accent": "#3e6c8e", "accent_hover": "#325a78",
        "accent_soft_bg": "#dce8ef", "accent_soft_fg": "#1f4258", "accent_soft_border": "#afc9d6",
        "success_fg": "#5b6b3f", "success_bg": "#e4e9d6", "success_border": "#8ca05e",
        "danger_fg": "#a2503f", "danger_bg": "#f2ddd6",
        "warning_bg": "#9c6b2e", "warning_hover": "#845a26", "warning_fg": "#ffffff",
        "purple_bg": "#6e5b8c", "purple_hover": "#5c4a78",
        "remarque_bg": "#f1e3cc", "remarque_border": "#b07c33", "remarque_fg": "#6b4a18",
        "btn_secondary_bg": "#e9e4d6", "btn_secondary_fg": "#3a392f", "btn_secondary_hover": "#ddd6c2",
        "chip_low_bg": "#e4e9d6", "chip_low_fg": "#4a5730",
        "chip_mid_bg": "#f1e3cc", "chip_mid_fg": "#6b4a18",
        "chip_high_bg": "#f2ddd6", "chip_high_fg": "#7c3b2d",
        "splitter_handle": "#dbd6c7",
        "statusbar_bg": "#23241f", "statusbar_fg": "#c9c5b4",
    },
    "dark": {
        "window_bg": "#181a16", "card_bg": "#20221d", "sidebar_bg": "#15170f",
        "text": "#ece8dd", "text_secondary": "#9c9a8c", "text_muted": "#6e6c5f",
        "border": "#34352c", "input_border": "#43443a",
        "accent": "#7fb0d6", "accent_hover": "#6699be",
        "accent_soft_bg": "#223441", "accent_soft_fg": "#cfe4f1", "accent_soft_border": "#3c566a",
        "success_fg": "#9cb06c", "success_bg": "#262b1b", "success_border": "#4c5a34",
        "danger_fg": "#e08979", "danger_bg": "#3b221d",
        "warning_bg": "#e0a85a", "warning_hover": "#c99247", "warning_fg": "#241a0d",
        "purple_bg": "#8c7fa6", "purple_hover": "#766693",
        "remarque_bg": "#3a2e1c", "remarque_border": "#e0a85a", "remarque_fg": "#f2d9ac",
        "btn_secondary_bg": "#2b2d24", "btn_secondary_fg": "#ece8dd", "btn_secondary_hover": "#363829",
        "chip_low_bg": "#262b1b", "chip_low_fg": "#b9ce8e",
        "chip_mid_bg": "#3a2e1c", "chip_mid_fg": "#f2d9ac",
        "chip_high_bg": "#3b221d", "chip_high_fg": "#f0afa0",
        "splitter_handle": "#20221d",
        "statusbar_bg": "#0e0f0c", "statusbar_fg": "#b9b6a6",
    },
}

# Thème actif — variable module ; basculée par MainWindow._toggle_theme().
CURRENT_THEME = "light"


def theme() -> dict:
    """Retourne le dictionnaire de couleurs du thème actuellement actif."""
    return THEMES[CURRENT_THEME]


def scrollbar_qss() -> str:
    """
    Feuille de style des ascenseurs : fins (10 px), sans flèches, poignée
    arrondie aux couleurs du thème — les scrollbars Windows par défaut
    (larges, grises) juraient avec le style « fiche technique » du reste.
    Appliquée aux deux zones scrollables (sidebar clients, panneau détail).
    """
    t = theme()
    return (
        "QScrollBar:vertical{background:transparent;width:10px;margin:2px 2px 2px 0;}"
        "QScrollBar:horizontal{background:transparent;height:10px;margin:0 2px 2px 2px;}"
        f"QScrollBar::handle:vertical{{background:{t['input_border']};border-radius:4px;min-height:32px;}}"
        f"QScrollBar::handle:horizontal{{background:{t['input_border']};border-radius:4px;min-width:32px;}}"
        f"QScrollBar::handle:hover{{background:{t['text_muted']};}}"
        "QScrollBar::add-line,QScrollBar::sub-line{width:0;height:0;}"
        "QScrollBar::add-page,QScrollBar::sub-page{background:transparent;}"
    )


def apply_titlebar_theme(widget):
    """
    Aligne la barre de titre Windows sur le thème actif : sans cet appel, une
    fenêtre en mode sombre garde une barre de titre blanche (détail qui trahit).
    Utilise DwmSetWindowAttribute avec l'attribut 20 (DWMWA_USE_IMMERSIVE_DARK_MODE,
    Windows 10 20H1+) puis 19 en repli pour les builds plus anciens.
    Silencieux (sans effet) hors Windows ou si l'appel échoue.
    """
    if sys.platform != "win32":
        return
    try:
        import ctypes
        hwnd  = int(widget.winId())
        value = ctypes.c_int(1 if CURRENT_THEME == "dark" else 0)
        for attribute in (20, 19):
            if ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, attribute, ctypes.byref(value), ctypes.sizeof(value)
            ) == 0:
                break
    except Exception:
        pass


class TitlebarThemer(QObject):
    """
    Filtre d'événements installé sur l'application (voir main()) : applique la
    barre de titre thémée à chaque fenêtre au moment où elle s'affiche —
    dialogues compris — sans avoir à modifier chaque classe de dialogue.
    """

    def eventFilter(self, obj, event):
        if (event.type() == QEvent.Type.Show
                and isinstance(obj, QWidget) and obj.isWindow()):
            apply_titlebar_theme(obj)
        return False


# Couleur de matière : chaque matière porte sa propre teinte (inspirée de sa
# couleur réelle — bleu acier « technique » pour le PTFE, ambre pour la résine
# PEEK) au lieu d'un accent bleu unique pour tout. Utilisée pour le sélecteur
# de matière, le bouton Calculer et la sélection dans la liste clients.
# "on" = couleur de texte lisible sur "solid" ; "soft_*" = variante discrète
# (badges, fonds légers). Une matière ajoutée par l'admin (au-delà de
# PTFE/PEEK) retombe sur DEFAULT_MATERIAL_ACCENT.
MATERIAL_ACCENTS = {
    "light": {
        "PTFE": {"solid": "#3e6c8e", "hover": "#325a78", "on": "#ffffff",
                 "soft_bg": "#dce8ef", "soft_fg": "#1f4258"},
        "PEEK": {"solid": "#9c6b2e", "hover": "#845a26", "on": "#ffffff",
                 "soft_bg": "#f1e3cc", "soft_fg": "#6b4a18"},
    },
    "dark": {
        "PTFE": {"solid": "#7fb0d6", "hover": "#6699be", "on": "#0f1720",
                 "soft_bg": "#223441", "soft_fg": "#cfe4f1"},
        "PEEK": {"solid": "#e0a85a", "hover": "#c99247", "on": "#241a0d",
                 "soft_bg": "#3a2e1c", "soft_fg": "#f2d9ac"},
    },
}
DEFAULT_MATERIAL_ACCENT = {
    "light": MATERIAL_ACCENTS["light"]["PTFE"],
    "dark":  MATERIAL_ACCENTS["dark"]["PTFE"],
}


def matiere_style(matiere: str) -> dict:
    """Retourne le jeu de couleurs (solid/hover/on/soft_bg/soft_fg) de la matière donnée."""
    return MATERIAL_ACCENTS.get(CURRENT_THEME, {}).get(matiere, DEFAULT_MATERIAL_ACCENT[CURRENT_THEME])


def print_palette(matiere: str | None = None) -> dict:
    """
    Palette fixe (toujours claire) pour les documents exportés (PDF, Excel) :
    indépendante du thème actif à l'écran, pour un rendu papier cohérent quel
    que soit le mode d'affichage. L'accent suit la matière du document quand
    elle est connue, sinon retombe sur le bleu acier générique.
    """
    light = THEMES["light"]
    accent = (
        MATERIAL_ACCENTS["light"].get(matiere, DEFAULT_MATERIAL_ACCENT["light"])["solid"]
        if matiere else light["accent"]
    )
    return {
        "accent":          accent,
        "ink":             light["text"],
        "ink_muted":       light["text_secondary"],
        "ink_faint":       light["text_muted"],
        "line":            light["border"],
        "stripe":          light["window_bg"],
        "success_bg":      light["success_bg"],
        "success_fg":      light["success_fg"],
        "success_border":  light["success_border"],
        "danger_fg":       light["danger_fg"],
        "remarque_bg":     light["card_bg"],
        "remarque_border": light["remarque_border"],
    }


def font_display(size: int, bold: bool = True) -> QFont:
    """
    Police des titres : Bahnschrift condensée (DIN-like, disponible nativement
    sous Windows 10+) — l'identité « fiche technique » repose sur ce contraste
    avec le Segoe UI du texte courant, pas seulement sur la couleur.
    """
    f = QFont("Bahnschrift", size, QFont.Weight.Bold if bold else QFont.Weight.Normal)
    f.setStretch(QFont.Stretch.SemiCondensed)
    return f


def font_mono(size: int, bold: bool = True) -> QFont:
    """Police tabulaire (Consolas) pour tout chiffre : prix, PL21, taux — lisible en colonne."""
    return QFont("Consolas", size, QFont.Weight.Bold if bold else QFont.Weight.Normal)


def make_theme_icon(kind: str, color: str, size: int = 18) -> QIcon:
    """
    Dessine une icône lune/soleil monochrome peinte à la main (QPainter), pour
    éviter le pictogramme emoji multicolore natif — seul contrôle de l'appli
    sans libellé texte, c'est le seul endroit qui a besoin d'une vraie icône.
    """
    pix = QPixmap(size, size)
    pix.fill(Qt.GlobalColor.transparent)
    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)
    p.setPen(Qt.PenStyle.NoPen)
    p.setBrush(QColor(color))
    cx = cy = size / 2
    if kind == "sun":
        r = size * 0.22
        p.drawEllipse(QPointF(cx, cy), r, r)
        pen = QPen(QColor(color))
        pen.setWidthF(1.4)
        pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        p.setPen(pen)
        for i in range(8):
            angle = i * math.pi / 4
            x1 = cx + math.cos(angle) * (r + 2.5)
            y1 = cy + math.sin(angle) * (r + 2.5)
            x2 = cx + math.cos(angle) * (r + 6)
            y2 = cy + math.sin(angle) * (r + 6)
            p.drawLine(QPointF(x1, y1), QPointF(x2, y2))
    else:  # "moon" : cercle plein moins un cercle décalé, en croissant
        r = size * 0.34
        full = QPainterPath()
        full.addEllipse(QPointF(cx, cy), r, r)
        cut = QPainterPath()
        cut.addEllipse(QPointF(cx + r * 0.55, cy - r * 0.35), r * 0.85, r * 0.85)
        p.drawPath(full.subtracted(cut))
    p.end()
    return QIcon(pix)


def make_dot_icon(color: str, size: int = 9) -> QIcon:
    """Petit disque plein coloré : indicateur de matière dans le sélecteur PTFE/PEEK."""
    pix = QPixmap(size, size)
    pix.fill(Qt.GlobalColor.transparent)
    p = QPainter(pix)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)
    p.setPen(Qt.PenStyle.NoPen)
    p.setBrush(QColor(color))
    p.drawEllipse(0, 0, size, size)
    p.end()
    return QIcon(pix)


def _client_initials(name: str) -> str:
    """Retourne 1-2 lettres majuscules représentant le client (pour l'avatar rond)."""
    parts = [p for p in name.strip().split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()


class StatCard(QFrame):
    """
    Petite carte de synthèse (intitulé + valeur) avec ombre portée.
    Utilisée dans ClientDetailPanel pour afficher augmentation / remise / prix final.
    variant="final" (prix final) et variant="danger" (remise) teintent seulement
    la valeur/le fond dans une couleur sémantique — pas de glow, la hiérarchie
    vient de la couleur, pas d'un effet lumineux.
    """

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self.setObjectName("statCard")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(3)
        self.lbl_title = QLabel(title)
        self.lbl_title.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        self.lbl_value = QLabel("—")
        self.lbl_value.setFont(font_mono(17))
        lay.addWidget(self.lbl_title)
        lay.addWidget(self.lbl_value)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(18)
        shadow.setOffset(0, 3)
        shadow.setColor(QColor(15, 23, 42, 35))
        self.setGraphicsEffect(shadow)
        self._num  = None   # dernière valeur numérique affichée (compteur animé)
        self._anim = None

    def set_value(self, text: str):
        if self._anim is not None:
            self._anim.stop()
            self._anim.deleteLater()
            self._anim = None
        self._num = None
        self.lbl_value.setText(text)

    def animate_value(self, target: float, fmt):
        """
        Effet « compteur » : la valeur défile de l'ancienne valeur affichée vers
        `target` en ~400 ms (easing OutCubic). `fmt` transforme le nombre en
        texte, ex. lambda v: f"{v:.0f} €". Premier affichage : défile depuis 0.
        """
        start = self._num if self._num is not None else 0.0
        self._num = target
        if self._anim is not None:
            self._anim.stop()
            self._anim.deleteLater()
            self._anim = None
        if start == target:
            self.lbl_value.setText(fmt(target))
            return
        anim = QVariantAnimation(self)
        anim.setStartValue(float(start))
        anim.setEndValue(float(target))
        anim.setDuration(420)
        anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        anim.valueChanged.connect(lambda v: self.lbl_value.setText(fmt(v)))
        anim.start()
        self._anim = anim

    def apply_theme(self, variant: str = "default"):
        t = theme()
        if variant == "final":
            bg, border, val_color = t["success_bg"], t["success_border"], t["success_fg"]
            title_color = t["success_fg"]
        elif variant == "danger":
            bg, border, val_color = t["card_bg"], t["border"], t["danger_fg"]
            title_color = t["text_secondary"]
        else:
            bg, border, val_color = t["card_bg"], t["border"], t["text"]
            title_color = t["text_secondary"]
        self.setStyleSheet(
            f"QFrame#statCard{{background:{bg};border:1px solid {border};border-radius:12px;}}"
        )
        self.lbl_title.setStyleSheet(f"color:{title_color};letter-spacing:1px;")
        self.lbl_value.setStyleSheet(f"color:{val_color};")
        eff = self.graphicsEffect()
        if isinstance(eff, QGraphicsDropShadowEffect):
            eff.setColor(QColor(15, 23, 42, 35))


class Toast(QLabel):
    """
    Notification éphémère : pastille qui glisse depuis le bas de la fenêtre
    (au-dessus de la barre de statut), reste ~1,5 s puis s'efface toute seule.
    Moins intrusive qu'une QMessageBox, plus visible qu'un message de statusbar.
    Une seule instance par fenêtre ; show_message() peut être rappelé à tout
    moment (les animations en cours sont simplement remplacées).
    """

    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
        self.hide()
        self._eff = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(self._eff)
        self._anims: list = []
        self._hide_timer = QTimer(self)
        self._hide_timer.setSingleShot(True)
        self._hide_timer.timeout.connect(self._fade_out)

    def show_message(self, text: str, duration_ms: int = 1600):
        t = theme()
        self.setText(text)
        # Taille calculée à la main (fontMetrics) : le padding QSS n'est pas
        # toujours pris en compte par sizeHint() sur un QLabel.
        fm = self.fontMetrics()
        w  = fm.horizontalAdvance(text) + 44
        h  = fm.height() + 16
        self.setFixedSize(w, h)
        self.setStyleSheet(
            f"background:{t['statusbar_bg']};color:{t['statusbar_fg']};"
            f"border-radius:{h // 2}px;"
        )
        parent = self.parentWidget()
        end   = QPoint((parent.width() - w) // 2, parent.height() - h - 48)
        start = QPoint(end.x(), end.y() + 16)

        for a in self._anims:
            a.stop()
            a.deleteLater()
        self._anims.clear()
        self._hide_timer.stop()

        self.move(start)
        self._eff.setOpacity(0.0)
        self.show()
        self.raise_()
        slide = QPropertyAnimation(self, b"pos", self)
        slide.setStartValue(start)
        slide.setEndValue(end)
        slide.setDuration(220)
        slide.setEasingCurve(QEasingCurve.Type.OutCubic)
        fade = QPropertyAnimation(self._eff, b"opacity", self)
        fade.setStartValue(0.0)
        fade.setEndValue(1.0)
        fade.setDuration(220)
        slide.start()
        fade.start()
        self._anims += [slide, fade]
        self._hide_timer.start(duration_ms)

    def _fade_out(self):
        fade = QPropertyAnimation(self._eff, b"opacity", self)
        fade.setStartValue(1.0)
        fade.setEndValue(0.0)
        fade.setDuration(260)
        fade.finished.connect(self.hide)
        fade.start()
        self._anims.append(fade)


# ─────────────────────────────────────────────────────────────────────────────
#  DIALOGUES ADMINISTRATEUR
# ─────────────────────────────────────────────────────────────────────────────

class EditClientDialog(QDialog):
    """
    Boîte de dialogue pour modifier les taux annuels et la remarque d'un client.
    Accessible via le bouton « ⚙ Modifier » du panneau de détail (mot de passe requis).

    Chaque année est représentée par un QDoubleSpinBox (valeur en %).
    On peut ajouter ou supprimer des années dynamiquement.
    Les modifications sont sauvegardées en JSON à la validation.
    """

    def __init__(self, matiere: str, client: str, parent=None):
        super().__init__(parent)
        self.matiere = matiere
        self.client  = client
        self.setWindowTitle(f"Modifier — {client} ({matiere})")
        self.setMinimumWidth(440)
        self.setModal(True)
        t = theme()
        self.setStyleSheet(
            f"QDialog{{background:{t['card_bg']};}}"
            f"QLabel{{color:{t['text']};}}QGroupBox{{color:{t['text']};}}"
        )
        self._build_ui()

    def _build_ui(self):
        """Construit l'interface : liste des années + champ remarque + boutons OK/Annuler."""
        t = theme()
        # On enroule tout dans un QScrollArea pour supporter de nombreuses années
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        container = QWidget()
        scroll.setWidget(container)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)
        layout = QVBoxLayout(container)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)

        titre = QLabel(f"Client : {self.client}  |  Matière : {self.matiere}")
        titre.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        titre.setStyleSheet(f"color:{t['accent']}; padding-bottom:4px;")
        layout.addWidget(titre)

        # Groupe taux annuels
        grp_taux = QGroupBox("Taux annuels d'augmentation")
        grp_taux.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        taux_layout = QVBoxLayout(grp_taux)
        self.taux_container = QWidget()
        self.taux_form = QFormLayout(self.taux_container)
        self.taux_form.setSpacing(8)
        self.taux_rows = {}  # dict { "2022": (QDoubleSpinBox, QWidget_ligne), ... }
        for annee in sorted(DATA["augmentations"][self.matiere][self.client].keys()):
            self._add_taux_row(annee, DATA["augmentations"][self.matiere][self.client][annee])
        taux_layout.addWidget(self.taux_container)
        btn_add = QPushButton("+ Ajouter une année")
        btn_add.setStyleSheet(
            f"QPushButton{{background:{t['accent_soft_bg']};color:{t['accent_soft_fg']};"
            f"border:1px solid {t['accent_soft_border']};border-radius:4px;padding:5px 10px;}}"
            f"QPushButton:hover{{background:{t['accent_soft_border']};}}"
        )
        btn_add.clicked.connect(self._ajouter_annee)
        taux_layout.addWidget(btn_add)
        layout.addWidget(grp_taux)

        # Groupe remarque (texte libre, optionnel)
        grp_rem = QGroupBox("Remarque (optionnel)")
        grp_rem.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        rem_layout = QVBoxLayout(grp_rem)
        self.input_remarque = QLineEdit()
        self.input_remarque.setPlaceholderText("Remarque spéciale…")
        self.input_remarque.setText(DATA["remarques"].get(self.client, ""))
        rem_layout.addWidget(self.input_remarque)
        layout.addWidget(grp_rem)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.button(QDialogButtonBox.StandardButton.Ok).setText("Enregistrer")
        btn_box.button(QDialogButtonBox.StandardButton.Ok).setStyleSheet(
            f"background:{t['accent']};color:white;padding:6px 16px;border-radius:4px;"
        )
        btn_box.accepted.connect(self._sauvegarder)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _add_taux_row(self, annee: str, valeur: float):
        """Ajoute une ligne « Année XXXX : [spinner %] [bouton ✕] » dans le formulaire."""
        t = theme()
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(8)
        spin = QDoubleSpinBox()
        pct = valeur * 100
        # Plage normale 0-100 %, élargie si besoin pour englober une valeur
        # existante hors bornes (ex: fichier édité à la main) : sans ça,
        # QDoubleSpinBox écrête silencieusement la valeur affichée, et
        # Enregistrer réécrirait alors une valeur tronquée sans avertissement.
        spin.setRange(min(0.0, pct), max(100.0, pct))
        spin.setDecimals(1)
        spin.setSuffix(" %")
        spin.setValue(pct)   # stocké en décimal, affiché en %
        spin.setFixedWidth(100)
        btn_suppr = QPushButton("✕")
        btn_suppr.setFixedSize(28, 28)
        btn_suppr.setStyleSheet(
            f"QPushButton{{background:{t['danger_bg']};color:{t['danger_fg']};border:none;"
            f"border-radius:4px;font-weight:bold;}}"
            f"QPushButton:hover{{background:{t['chip_high_bg']};}}"
        )
        # La lambda capture 'annee' par défaut pour éviter la capture par référence
        btn_suppr.clicked.connect(lambda _, a=annee: self._supprimer_annee(a))
        row_layout.addWidget(spin)
        row_layout.addWidget(btn_suppr)
        row_layout.addStretch()
        self.taux_form.addRow(f"Année {annee} :", row_widget)
        self.taux_rows[annee] = (spin, row_widget)

    def _ajouter_annee(self):
        """Demande une année à l'utilisateur et ajoute une ligne vide (taux 0 %)."""
        annee, ok = QInputDialog.getText(self, "Nouvelle année", "Année (ex: 2027) :")
        if not ok or not annee.strip():
            return
        annee = annee.strip()
        if not annee.isdigit() or len(annee) != 4:
            QMessageBox.warning(self, "Erreur", "L'année doit être un nombre à 4 chiffres.")
            return
        if annee in self.taux_rows:
            QMessageBox.warning(self, "Erreur", f"L'année {annee} existe déjà.")
            return
        self._add_taux_row(annee, 0.0)

    def _supprimer_annee(self, annee: str):
        """Supprime une ligne d'année du formulaire (minimum 1 année obligatoire)."""
        if len(self.taux_rows) <= 1:
            QMessageBox.warning(self, "Erreur", "Il faut au moins une année.")
            return
        _, row_widget = self.taux_rows.pop(annee)
        # Trouver et supprimer la ligne correspondante dans le QFormLayout
        for i in range(self.taux_form.rowCount()):
            item = self.taux_form.itemAt(i, QFormLayout.ItemRole.FieldRole)
            if item and item.widget() == row_widget:
                self.taux_form.removeRow(i)
                break

    def _sauvegarder(self):
        """
        Valide le formulaire, met à jour DATA en mémoire et sur disque,
        trace l'action dans le journal d'audit, puis ferme la boîte de dialogue.
        """
        # Reconvertir les valeurs % → décimal et arrondir à 4 décimales
        nouveaux_taux = {
            a: round(spin.value() / 100, 4)
            for a, (spin, _) in self.taux_rows.items()
        }
        DATA["augmentations"][self.matiere][self.client] = nouveaux_taux
        # Mettre à jour ou supprimer la remarque
        rem = self.input_remarque.text().strip()
        if rem:
            DATA["remarques"][self.client] = rem
        elif self.client in DATA["remarques"]:
            del DATA["remarques"][self.client]
        # log_audit() persiste DATA : pas besoin d'un save_data() séparé ici.
        log_audit("edit_client", self.matiere, self.client, f"Taux: {nouveaux_taux}")
        self.accept()


class AddClientDialog(QDialog):
    """
    Boîte de dialogue pour créer un nouveau client.
    Le client est ajouté à TOUTES les matières existantes avec des taux à 0 %
    (à affiner ensuite via EditClientDialog).
    La remise saisie s'applique également à toutes les matières.
    """

    def __init__(self, matiere: str, parent=None):
        super().__init__(parent)
        self.matiere    = matiere
        self._nom_final = ""   # nom retourné après validation (voir get_nom())
        self.setWindowTitle("Ajouter un client")
        self.setFixedWidth(400)
        self.setModal(True)
        t = theme()
        self.setStyleSheet(
            f"QDialog{{background:{t['card_bg']};}}"
            f"QLabel{{color:{t['text']};}}QGroupBox{{color:{t['text']};}}"
        )
        layout = QVBoxLayout(self)
        layout.setSpacing(14)
        layout.setContentsMargins(20, 20, 20, 20)

        titre = QLabel("Nouveau client")
        titre.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        titre.setStyleSheet(f"color:{t['accent']};")
        layout.addWidget(titre)

        grp = QGroupBox("Informations")
        form = QFormLayout(grp)
        self.input_nom = QLineEdit()
        self.input_nom.setPlaceholderText("Nom du client")
        self.input_nom.setFont(QFont("Segoe UI", 11))
        form.addRow("Nom :", self.input_nom)
        self.spin_remise = QDoubleSpinBox()
        self.spin_remise.setRange(0, 99)
        self.spin_remise.setDecimals(1)
        self.spin_remise.setSuffix(" %")
        self.spin_remise.setValue(60.0)
        form.addRow("Remise (toutes matières) :", self.spin_remise)
        layout.addWidget(grp)

        info = QLabel(
            "ℹ Les taux annuels seront initialisés à 0 % et pourront être "
            "modifiés via ⚙ Modifier."
        )
        info.setWordWrap(True)
        info.setStyleSheet(f"color:{t['text_secondary']};font-size:9pt;")
        layout.addWidget(info)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Ajouter")
        btns.button(QDialogButtonBox.StandardButton.Ok).setStyleSheet(
            f"background:{t['accent']};color:white;padding:5px 14px;border-radius:4px;"
        )
        btns.accepted.connect(self._valider)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _valider(self):
        """
        Vérifie que le nom est non vide et n'existe pas déjà,
        puis crée le client dans toutes les matières.
        """
        nom = self.input_nom.text().strip()
        if not nom:
            QMessageBox.warning(self, "Erreur", "Le nom est requis.")
            return
        if any(nom in mats for mats in DATA["augmentations"].values()):
            QMessageBox.warning(self, "Erreur", f"'{nom}' existe déjà.")
            return
        remise = -round(self.spin_remise.value() / 100, 4)
        # Années : union de toutes les années utilisées par les clients existants
        # (toutes matières confondues), plutôt que celles du seul premier client
        # rencontré, qui pourrait avoir un jeu d'années non représentatif.
        annees = {
            annee
            for mats in DATA["augmentations"].values()
            for taux in mats.values()
            for annee in taux
        }
        taux_vides = {annee: 0.0 for annee in annees}
        for mat in DATA["augmentations"]:
            DATA["augmentations"][mat][nom] = copy.deepcopy(taux_vides)
            # setdefault plutôt qu'un indexage nu : robuste même si 'remises'
            # ne contenait pas encore cette matière (fichier édité à la main).
            DATA["remises"].setdefault(mat, {})[nom] = remise
        # log_audit() persiste DATA : pas besoin d'un save_data() séparé ici.
        log_audit("add_client", self.matiere, nom)
        self._nom_final = nom
        self.accept()

    def get_nom(self) -> str:
        """Retourne le nom du client créé (disponible après exec() == Accepted)."""
        return self._nom_final


class ChangePasswordDialog(QDialog):
    """
    Boîte de dialogue pour changer le mot de passe administrateur.
    Le nouveau mot de passe doit faire au moins 4 caractères et être saisi deux fois.
    Un hash salé (PBKDF2-HMAC-SHA256) est stocké dans DATA, jamais le mot de
    passe en clair (voir _hash_password).
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Changer le mot de passe administrateur")
        self.setFixedWidth(360)
        self.setModal(True)
        t = theme()
        self.setStyleSheet(
            f"QDialog{{background:{t['card_bg']};}}"
            f"QLabel{{color:{t['text']};}}"
        )
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        titre = QLabel("Nouveau mot de passe")
        titre.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        titre.setStyleSheet(f"color:{t['accent']};")
        layout.addWidget(titre)

        form = QFormLayout()
        self.inp_new  = QLineEdit()
        self.inp_new.setEchoMode(QLineEdit.EchoMode.Password)
        self.inp_conf = QLineEdit()
        self.inp_conf.setEchoMode(QLineEdit.EchoMode.Password)
        form.addRow("Nouveau :",   self.inp_new)
        form.addRow("Confirmer :", self.inp_conf)
        layout.addLayout(form)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Changer")
        btns.button(QDialogButtonBox.StandardButton.Ok).setStyleSheet(
            f"background:{t['accent']};color:white;padding:5px 14px;border-radius:4px;"
        )
        btns.accepted.connect(self._valider)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _valider(self):
        new, conf = self.inp_new.text(), self.inp_conf.text()
        if len(new) < 4:
            QMessageBox.warning(self, "Erreur", "Minimum 4 caractères.")
            return
        if new != conf:
            QMessageBox.warning(self, "Erreur", "Les mots de passe ne correspondent pas.")
            return
        set_admin_hash(_hash_password(new))
        QMessageBox.information(self, "Succès", "Mot de passe modifié.")
        self.accept()


class AuditLogDialog(QDialog):
    """
    Affiche le journal complet des modifications administratives (lecture seule).
    Les entrées sont affichées du plus récent au plus ancien.
    Accessible via Administration → Journal des modifications…
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Journal des modifications")
        self.setMinimumSize(750, 480)
        t = theme()
        self.setStyleSheet(f"QDialog{{background:{t['card_bg']};}}QLabel{{color:{t['text']};}}")
        layout = QVBoxLayout(self)

        log = list(reversed(DATA.get("audit_log", [])))
        table = QTableWidget(len(log), 5)
        table.setHorizontalHeaderLabels(["Date/Heure", "Action", "Matière", "Client", "Détails"])
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        for i, e in enumerate(log):
            for j, key in enumerate(["ts", "action", "matiere", "client", "details"]):
                table.setItem(i, j, QTableWidgetItem(str(e.get(key, ""))))
        layout.addWidget(table)

        btn = QPushButton("Fermer")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)


class HistoriqueDialog(QDialog):
    """
    Affiche l'historique des calculs (PL21 utilisé, client, prix obtenu).
    Les entrées sont affichées du plus récent au plus ancien.
    Permet de vider entièrement l'historique (irréversible).
    Accessible via Fichier → Historique des calculs…
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Historique des calculs")
        self.setMinimumSize(700, 480)
        t = theme()
        self.setStyleSheet(f"QDialog{{background:{t['card_bg']};}}QLabel{{color:{t['text']};}}")
        layout = QVBoxLayout(self)

        hbox = QHBoxLayout()
        btn_clear = QPushButton("Vider l'historique")
        btn_clear.setStyleSheet(
            f"background:{t['danger_bg']};color:{t['danger_fg']};border-radius:4px;padding:4px 12px;"
        )
        btn_clear.clicked.connect(self._vider)
        hbox.addStretch()
        hbox.addWidget(btn_clear)
        layout.addLayout(hbox)

        historique = list(reversed(DATA.get("historique", [])))
        self.table = QTableWidget(len(historique), 5)
        self.table.setHorizontalHeaderLabels(
            ["Date/Heure", "Matière", "Client", "PL21 (€)", "Prix final (€)"]
        )
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        font_vert = QFont("Segoe UI", 10)
        font_vert.setBold(True)
        for i, e in enumerate(historique):
            for j, key in enumerate(["ts", "matiere", "client", "pl21", "prix_final"]):
                item = QTableWidgetItem(str(e.get(key, "")))
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignCenter
                )
                if j == 4:   # colonne prix final en vert
                    item.setForeground(QColor(t["success_fg"]))
                    item.setFont(font_vert)
                self.table.setItem(i, j, item)
        layout.addWidget(self.table)

        btn = QPushButton("Fermer")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)

    def _vider(self):
        """Efface tout l'historique après confirmation de l'utilisateur."""
        if QMessageBox.question(
            self, "Confirmer", "Vider tout l'historique ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            DATA["historique"] = []
            save_data(DATA)
            self.table.setRowCount(0)


# ─────────────────────────────────────────────────────────────────────────────
#  ITEM DE LA LISTE CLIENTS — sidebar gauche
# ─────────────────────────────────────────────────────────────────────────────

class ClientListItem(QFrame):
    """
    Widget cliquable représentant un client dans la sidebar.

    Affiche, sur une seule ligne (avatar | nom | prix aligné à droite) —
    une liste dense façon relevé technique plutôt que des cartes espacées :
      • un avatar rond avec les initiales du client
      • le nom du client (semi-gras)
      • le prix final calculé (chiffres tabulaires) ou « — » si pas encore calculé

    États visuels :
      • inactif : fond transparent, avatar neutre, survol coloré selon le thème
      • actif   : liseré + avatar dans la couleur de la matière du client,
                  fond légèrement teinté (pas un aplat saturé) — la couleur
                  reste un signal (matière + sélection), pas une décoration.

    on_click : callable appelé avec le nom du client quand l'utilisateur clique.
    """

    def __init__(self, client: str, matiere: str, on_click, parent=None):
        super().__init__(parent)
        self.client     = client
        self.matiere    = matiere
        self._on_click  = on_click
        self._active    = False
        self._has_price = False

        self.setFixedHeight(42)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFrameShape(QFrame.Shape.NoFrame)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(13, 4, 14, 4)
        layout.setSpacing(10)

        # Avatar rond : initiales, neutre au repos, coloré (matière) une fois sélectionné
        self.lbl_avatar = QLabel(_client_initials(client))
        self.lbl_avatar.setFixedSize(24, 24)
        self.lbl_avatar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_avatar.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        layout.addWidget(self.lbl_avatar)

        # Nom et prix sur une seule ligne (nom à gauche, prix aligné à droite) —
        # plus dense que l'ancien nom/prix empilés sur deux lignes.
        self.lbl_nom = QLabel(client)
        self.lbl_nom.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
        self.lbl_prix = QLabel("—")
        self.lbl_prix.setFont(font_mono(9))
        self.lbl_prix.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(self.lbl_nom, 1)
        layout.addWidget(self.lbl_prix)

        self.apply_theme()

    def set_prix(self, prix_final: int):
        """Affiche le prix final calculé (appelé depuis MainWindow._calculer_tout)."""
        self._has_price = True
        self.lbl_prix.setText(f"{prix_final} €")
        self._restyle_price()

    def clear_prix(self):
        """Remet « — » quand l'utilisateur clique sur Réinitialiser."""
        self._has_price = False
        self.lbl_prix.setText("—")
        self._restyle_price()

    def set_active(self, active: bool):
        """Bascule l'état sélectionné/non sélectionné et met à jour les couleurs."""
        self._active = active
        self.apply_theme()

    def apply_theme(self):
        """Réapplique les couleurs de fond/texte/avatar selon le thème et l'état sélectionné."""
        t  = theme()
        ms = matiere_style(self.matiere)
        if self._active:
            # Fond teinté dans la couleur de la matière du client (PTFE/PEEK/…) :
            # la sélection annonce la matière au lieu d'un bleu générique — sans
            # liseré, juste le fond et l'avatar coloré.
            self.setStyleSheet(f"QFrame{{background:{ms['soft_bg']};border:none;}}")
            self.lbl_nom.setStyleSheet(f"color:{ms['soft_fg']};")
            self.lbl_avatar.setStyleSheet(
                f"background:{ms['solid']};color:{ms['on']};border-radius:12px;"
            )
        else:
            self.setStyleSheet(
                "QFrame{background:transparent;border:none;}"
                f"QFrame:hover{{background:{t['btn_secondary_bg']};}}"
            )
            self.lbl_nom.setStyleSheet(f"color:{t['text']};")
            self.lbl_avatar.setStyleSheet(
                f"background:{t['btn_secondary_bg']};color:{t['text_secondary']};border-radius:12px;"
            )
        self._restyle_price()

    def _restyle_price(self):
        t = theme()
        if self._active:
            color = matiere_style(self.matiere)["soft_fg"]
        else:
            color = t["success_fg"] if self._has_price else t["text_muted"]
        weight = "font-weight:bold;" if self._has_price else ""
        self.lbl_prix.setStyleSheet(f"color:{color};{weight}")

    def mousePressEvent(self, event):
        """Déclenche le callback de sélection sur un clic gauche uniquement."""
        if event.button() == Qt.MouseButton.LeftButton:
            self._on_click(self.client)


# ─────────────────────────────────────────────────────────────────────────────
#  CASCADE DE CALCUL — remplace le tableau Année/Taux/Prix
# ─────────────────────────────────────────────────────────────────────────────

class WaterfallChart(QWidget):
    """
    Visualise la construction du prix comme une cascade : PL21, puis chaque
    année applique sa hausse sur le prix précédent, puis la remise retire une
    part du total — la formule prix_final = ceil(PL21 × ∏(1+taux) × (1+remise))
    rendue lisible d'un coup d'œil plutôt qu'en lignes de tableau.

    set_steps() attend une liste d'étapes, chacune un dict :
      {"kind": "start"|"increase"|"final", "name": str, "value": float,
       "prev_value": float (absent pour "start"),
       "delta_text": str (absent pour "start"), "value_text": str}
    Tant qu'aucun PL21 n'est saisi, l'appelant fournit une cascade indexée sur
    une base 100 (la forme reste juste) avec value_text="—".

    Animation : avec set_steps(..., animate=True), les colonnes « poussent »
    depuis la ligne de base une à une, en léger décalé — le calcul se construit
    sous les yeux de l'utilisateur au lieu d'apparaître d'un bloc.
    Survol : la colonne sous la souris est mise en surbrillance et une
    infobulle détaille l'étape (prix avant → taux → prix après).
    """

    # Part de la timeline totale occupée par la pousse d'UNE colonne ; le reste
    # (1 - COL_SPAN) est réparti en décalages de démarrage entre colonnes.
    COL_SPAN = 0.45

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(270)
        self.setMouseTracking(True)   # survol des colonnes sans clic
        self._steps: list[dict] = []
        self._matiere: str | None = None
        self._hover_col = -1
        self._progress  = 1.0         # progression de l'animation (1 = tout affiché)
        self._anim = QVariantAnimation(self)
        self._anim.setStartValue(0.0)
        self._anim.setEndValue(1.0)
        self._anim.valueChanged.connect(self._on_anim_tick)
        # Polices constantes du graphique : créées une seule fois plutôt qu'à
        # chaque paintEvent (~60 fois/seconde pendant la pousse animée d'une
        # cascade) — QFont reste identique tant que le widget vit.
        self._font_delta = QFont("Segoe UI", 10, QFont.Weight.Bold)
        self._font_value = QFont("Consolas", 12, QFont.Weight.Bold)
        self._font_name  = QFont("Segoe UI", 9, QFont.Weight.DemiBold)

    def _on_anim_tick(self, value):
        self._progress = float(value)
        self.update()

    def set_steps(self, steps: list[dict], matiere: str, animate: bool = False):
        self._steps = steps
        self._matiere = matiere
        self._hover_col = -1
        self._anim.stop()
        if animate and steps:
            self._anim.setDuration(450 + 80 * len(steps))
            self._progress = 0.0
            self._anim.start()
        else:
            self._progress = 1.0
        self.update()

    def apply_theme(self):
        self.update()

    # ── Géométrie partagée entre paintEvent et la détection de survol ────────
    def _geometry(self):
        """Retourne (x0, col_w, gap, base_y, margin_top, plot_h, max_val)."""
        n = len(self._steps)
        margin_top, margin_bottom = 34, 56
        plot_h = max(self.height() - margin_top - margin_bottom, 10)
        gap = 20
        col_w = min(76.0, max((self.width() - gap * (n - 1)) / n, 10.0))
        total_w = col_w * n + gap * (n - 1)
        x0 = (self.width() - total_w) / 2
        base_y = self.height() - margin_bottom
        max_val = max(s["value"] for s in self._steps) or 1.0
        return x0, col_w, gap, base_y, margin_top, plot_h, max_val

    def _col_progress(self, i: int) -> float:
        """Progression (0→1, easing OutCubic) de la colonne i, démarrage décalé."""
        if self._progress >= 1.0:
            return 1.0
        n = len(self._steps)
        start = (i / max(n - 1, 1)) * (1 - self.COL_SPAN)
        x = (self._progress - start) / self.COL_SPAN
        x = max(0.0, min(1.0, x))
        return 1 - (1 - x) ** 3

    # ── Survol : surbrillance + infobulle détaillée ──────────────────────────
    def _col_at(self, x: float, y: float) -> int:
        """Indice de la colonne sous (x, y), ou -1 si aucune."""
        if not self._steps:
            return -1
        x0, col_w, gap, base_y, margin_top, _, _ = self._geometry()
        if not (margin_top - 28 <= y <= base_y + 50):
            return -1
        for i in range(len(self._steps)):
            cx = x0 + i * (col_w + gap)
            if cx <= x <= cx + col_w:
                return i
        return -1

    def _tooltip_for(self, i: int) -> str:
        s = self._steps[i]
        if s["kind"] == "start":
            return f"PL21 (prix catalogue 2021) : {s['value_text']}"
        avant = self._steps[i - 1]["value_text"]
        if s["kind"] == "increase":
            return f"{s['name']} : {s['delta_text']}\n{avant}  →  {s['value_text']}"
        return f"Remise {s['delta_text']}\n{avant}  →  {s['value_text']}"

    def mouseMoveEvent(self, event):
        pos = event.position()
        col = self._col_at(pos.x(), pos.y())
        if col != self._hover_col:
            self._hover_col = col
            self.update()
        if col >= 0:
            QToolTip.showText(event.globalPosition().toPoint(), self._tooltip_for(col), self)
        else:
            QToolTip.hideText()

    def leaveEvent(self, _event):
        if self._hover_col != -1:
            self._hover_col = -1
            self.update()

    def paintEvent(self, _event):
        if not self._steps:
            return
        t  = theme()
        ms = matiere_style(self._matiere)
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        x0, col_w, gap, base_y, margin_top, plot_h, max_val = self._geometry()

        def y_of(v):
            return margin_top + plot_h * (1 - min(v, max_val) / max_val)

        font_delta = QFont("Segoe UI", 10, QFont.Weight.Bold)
        font_value = QFont("Consolas", 12, QFont.Weight.Bold)
        font_name  = QFont("Segoe UI", 9, QFont.Weight.DemiBold)
        hover_tint = (QColor(255, 255, 255, 46) if CURRENT_THEME == "dark"
                      else QColor(35, 36, 31, 22))

        for i, step in enumerate(self._steps):
            prog = self._col_progress(i)
            if prog <= 0.0:
                continue
            x = x0 + i * (col_w + gap)
            top_y  = y_of(step["value"])
            prev_y = y_of(step["prev_value"]) if "prev_value" in step else top_y
            col_top = min(top_y, prev_y) - 26   # inclut le libellé de delta au-dessus

            # Pendant l'animation, la colonne « pousse » depuis la ligne de base :
            # dessin limité à un rectangle qui grandit vers le haut, en fondu.
            p.save()
            p.setOpacity(prog)
            if prog < 1.0:
                clip_h = (base_y - col_top) * prog
                p.setClipRect(QRectF(x - gap - 14, base_y - clip_h,
                                     col_w + 2 * gap + 28, clip_h + 1))

            if step["kind"] == "start":
                p.setPen(Qt.PenStyle.NoPen)
                p.setBrush(QColor(t["text_muted"]))
                p.drawRect(QRectF(x, top_y, col_w, base_y - top_y))

            elif step["kind"] == "increase":
                p.setPen(Qt.PenStyle.NoPen)
                p.setBrush(QColor(t["border"]))
                p.drawRect(QRectF(x, prev_y, col_w, base_y - prev_y))
                p.setBrush(QColor(ms["solid"]))
                p.drawRect(QRectF(x, top_y, col_w, max(prev_y - top_y, 1.5)))
                pen = QPen(QColor(t["text_muted"]))
                pen.setStyle(Qt.PenStyle.DashLine)
                p.setPen(pen)
                p.drawLine(QPointF(x - gap, prev_y), QPointF(x, prev_y))
                p.setPen(QColor(ms["solid"]))
                p.setFont(font_delta)
                p.drawText(QRectF(x - 16, top_y - 24, col_w + 32, 20),
                           Qt.AlignmentFlag.AlignCenter, step["delta_text"])

            else:  # "final" — la remise retire une part du dernier prix cumulé
                ghost = QRectF(x, prev_y, col_w, max(top_y - prev_y, 1.5))
                p.setPen(Qt.PenStyle.NoPen)
                p.setBrush(QColor(t["danger_bg"]))
                p.drawRect(ghost)
                p.setBrush(QBrush(QColor(t["danger_fg"]), Qt.BrushStyle.BDiagPattern))
                p.drawRect(ghost)
                p.setPen(QPen(QColor(t["success_border"]), 1.3))
                p.setBrush(QColor(t["success_bg"]))
                p.drawRect(QRectF(x, top_y, col_w, base_y - top_y))
                pen = QPen(QColor(t["text_muted"]))
                pen.setStyle(Qt.PenStyle.DashLine)
                p.setPen(pen)
                p.drawLine(QPointF(x - gap, prev_y), QPointF(x, prev_y))
                p.setPen(QColor(t["danger_fg"]))
                p.setFont(font_delta)
                p.drawText(QRectF(x - 16, prev_y - 24, col_w + 32, 20),
                           Qt.AlignmentFlag.AlignCenter, step["delta_text"])

            # Surbrillance de la colonne survolée
            if i == self._hover_col:
                p.setPen(Qt.PenStyle.NoPen)
                p.setBrush(hover_tint)
                p.drawRect(QRectF(x, min(top_y, prev_y), col_w,
                                  base_y - min(top_y, prev_y)))
            p.restore()

            # Libellés sous la ligne de base (valeur + nom), en fondu avec la colonne
            p.setOpacity(prog)
            value_color = t["success_fg"] if step["kind"] == "final" else t["text"]
            p.setPen(QColor(value_color))
            p.setFont(font_value)
            p.drawText(QRectF(x - 16, base_y + 6, col_w + 32, 22),
                       Qt.AlignmentFlag.AlignCenter, step["value_text"])
            p.setPen(QColor(t["text_muted"]))
            p.setFont(font_name)
            p.drawText(QRectF(x - 16, base_y + 28, col_w + 32, 18),
                       Qt.AlignmentFlag.AlignCenter, step["name"])
            p.setOpacity(1.0)


# ─────────────────────────────────────────────────────────────────────────────
#  PANNEAU DÉTAIL CLIENT — zone de droite
# ─────────────────────────────────────────────────────────────────────────────

class ClientDetailPanel(QScrollArea):
    """
    Panneau scrollable affiché à droite du splitter.
    Montre le détail de calcul pour le client sélectionné dans la sidebar.

    Contenu :
      • En-tête : nom du client + bouton « PDF » + bouton « ⚙ Modifier »
      • Cascade de calcul (WaterfallChart) : PL21 → hausses annuelles → remise
      • Lien « ✎ Modifier la remise » (ouvert sans mot de passe)
      • Cartes de synthèse + bouton copier le prix final
      • Zone remarque (si une remarque existe pour ce client)

    Points d'entrée principaux :
      load_client(client, matiere) — change le client affiché
      update_pl21(pl21)           — déclenche le calcul et met à jour l'affichage
    """

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self._client   = None    # nom du client actuellement affiché
        self._matiere  = None    # matière actuellement affichée
        self._pl21     = None    # dernière valeur PL21 calculée (None = pas encore calculé)
        self._last_res = None    # dernier résultat de calculer_prix() (pour le bouton copier)
        self._fade_anim = None   # animation de fondu au changement de client

        self.setWidgetResizable(True)
        self.setFrameShape(QFrame.Shape.NoFrame)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        container = QWidget()
        self.setWidget(container)
        self._layout = QVBoxLayout(container)
        self._layout.setSpacing(16)
        self._layout.setContentsMargins(24, 20, 24, 20)

        # État vide : un client est normalement auto-sélectionné, donc ce
        # panneau ne s'affiche que si la matière ne contient plus aucun client.
        self._placeholder = QLabel(
            "Aucun client dans cette matière.\n\n"
            "Menu Administration → « Ajouter un client… » pour en créer un."
        )
        self._placeholder.setFont(QFont("Segoe UI", 12))
        self._placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._placeholder.setWordWrap(True)
        self._layout.addWidget(self._placeholder)
        self._layout.addStretch()

        # Contenu principal (caché jusqu'à la première sélection)
        self._content = QWidget()
        self._content.hide()
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setSpacing(16)
        self._content_layout.setContentsMargins(0, 0, 0, 0)
        # Inséré en position 0 pour apparaître avant le placeholder
        self._layout.insertWidget(0, self._content)

        self._build_content()

    def _build_content(self):
        """
        Construit les widgets permanents du panneau de détail.
        Cette méthode n'est appelée qu'une seule fois à l'initialisation.
        Le contenu variable (la cascade) est (re)créé dans _refresh_chart().
        Les couleurs sont appliquées séparément par apply_theme().
        """
        lay = self._content_layout

        # ── En-tête ──────────────────────────────────────────────────────────
        header_row = QHBoxLayout()
        header_row.setSpacing(10)
        self.lbl_heading = QLabel("")
        self.lbl_heading.setFont(font_display(18))

        # Petit badge rappelant la matière du client affiché (PTFE/PEEK/…)
        self.lbl_matiere_chip = QLabel("")
        self.lbl_matiere_chip.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))

        # Bouton PDF (violet) — export pour ce client uniquement
        self.btn_pdf = QPushButton("PDF")
        self.btn_pdf.setFixedHeight(32)
        self.btn_pdf.setFont(QFont("Segoe UI", 10))
        self.btn_pdf.clicked.connect(self._export_pdf_client)

        # Bouton Modifier (orange) — ouvre EditClientDialog après vérification mot de passe
        self.btn_admin = QPushButton("⚙ Modifier")
        self.btn_admin.setFixedHeight(32)
        self.btn_admin.setFont(QFont("Segoe UI", 10))
        self.btn_admin.clicked.connect(self._ouvrir_edition)

        header_row.addWidget(self.lbl_heading)
        header_row.addWidget(self.lbl_matiere_chip)
        header_row.addStretch()
        header_row.addWidget(self.btn_pdf)
        header_row.addWidget(self.btn_admin)
        lay.addLayout(header_row)

        # ── Cartes de synthèse (augmentation totale / remise / prix final) ────
        stats_row = QHBoxLayout()
        stats_row.setSpacing(12)
        self.card_total  = StatCard("AUGMENTATION TOTALE")
        self.card_remise = StatCard("REMISE")
        self.card_final  = StatCard("PRIX FINAL")
        # Le prix final pèse plus lourd visuellement (1.3x) : c'est la seule
        # valeur qui compte vraiment pour l'utilisateur au final.
        # Alignement en haut : la colonne "remise" (carte + lien) est plus
        # haute que les deux autres cartes, qui ne doivent pas s'étirer pour
        # combler la différence.
        stats_row.addWidget(self.card_total, 10, Qt.AlignmentFlag.AlignTop)

        remise_col = QVBoxLayout()
        remise_col.setSpacing(4)
        remise_col.addWidget(self.card_remise)
        self.btn_edit_remise = QPushButton("✎ Modifier la remise")
        self.btn_edit_remise.setFlat(True)
        self.btn_edit_remise.setFont(QFont("Segoe UI", 8))
        self.btn_edit_remise.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_edit_remise.clicked.connect(self._ouvrir_remise_dialog)
        remise_col.addWidget(self.btn_edit_remise, alignment=Qt.AlignmentFlag.AlignHCenter)
        stats_row.addLayout(remise_col, 10)

        stats_row.addWidget(self.card_final, 13, Qt.AlignmentFlag.AlignTop)
        lay.addLayout(stats_row)

        # ── Bouton copier (visible/actif selon l'état) ─────────────────────
        copy_row = QHBoxLayout()
        copy_row.addStretch()
        self.btn_copier = QPushButton("Copier le prix final")
        self.btn_copier.setFixedHeight(34)
        self.btn_copier.setFont(QFont("Segoe UI", 10))
        self.btn_copier.clicked.connect(self._copier_prix)
        copy_row.addWidget(self.btn_copier)
        lay.addLayout(copy_row)

        # ── Titre + carte de la cascade de calcul ───────────────────────────────
        self.lbl_section_detail = QLabel("CONSTRUCTION DU PRIX")
        self.lbl_section_detail.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        lay.addWidget(self.lbl_section_detail)

        # Le contenu (la cascade) est mis à jour par _refresh_chart()
        self.result_frame = QFrame()
        self.result_frame.setObjectName("resultCard")
        shadow = QGraphicsDropShadowEffect(self.result_frame)
        shadow.setBlurRadius(20)
        shadow.setOffset(0, 4)
        shadow.setColor(QColor(15, 23, 42, 30))
        self.result_frame.setGraphicsEffect(shadow)
        result_outer = QVBoxLayout(self.result_frame)
        result_outer.setContentsMargins(20, 20, 20, 20)
        self.waterfall = WaterfallChart()
        result_outer.addWidget(self.waterfall)
        lay.addWidget(self.result_frame)

        # ── Zone de remarque (cachée si pas de remarque) ──────────────────────
        self.lbl_remarque = QLabel()
        self.lbl_remarque.setWordWrap(True)
        lay.addWidget(self.lbl_remarque)
        lay.addStretch()

        self.apply_theme()

    def apply_theme(self):
        """
        Réapplique toutes les couleurs du panneau selon le thème actif.
        Repeint aussi la cascade (couleurs dépendantes du thème/de la matière).
        Appelé à l'initialisation et à chaque bascule clair/sombre.
        """
        t = theme()
        self.setStyleSheet(scrollbar_qss())   # ascenseurs fins aux couleurs du thème
        self._placeholder.setStyleSheet(f"color:{t['text_muted']};")
        self.lbl_heading.setStyleSheet(f"color:{t['text']};")
        # Boutons d'action discrets (bordure + texte muted) : la couleur reste
        # réservée à ce qui porte un sens (matière, remise, prix final), pas
        # à des actions secondaires comme exporter ou éditer.
        ghost_btn = (
            f"QPushButton{{background:transparent;color:{t['text_secondary']};"
            f"border:1px solid {t['border']};border-radius:6px;padding:3px 12px;}}"
            f"QPushButton:hover{{background:{t['btn_secondary_hover']};color:{t['text']};}}"
        )
        self.btn_pdf.setStyleSheet(ghost_btn)
        self.btn_admin.setStyleSheet(ghost_btn)
        self.btn_edit_remise.setStyleSheet(
            f"QPushButton{{color:{t['text_muted']};background:transparent;border:none;padding:4px 6px;}}"
            f"QPushButton:hover{{color:{t['text']};}}"
        )
        self.btn_copier.setStyleSheet(
            f"QPushButton{{background:{t['accent_soft_bg']};color:{t['accent_soft_fg']};"
            f"border-radius:6px;border:1px solid {t['accent_soft_border']};padding:0 14px;}}"
            f"QPushButton:hover{{background:{t['accent']};color:white;}}"
        )
        self.card_total.apply_theme("default")
        self.card_remise.apply_theme("danger")
        self.card_final.apply_theme("final")
        self.lbl_section_detail.setStyleSheet(f"color:{t['text_secondary']};letter-spacing:1px;")
        self.result_frame.setStyleSheet(
            f"QFrame#resultCard{{background:{t['card_bg']};border:1px solid {t['border']};"
            f"border-radius:14px;}}"
        )
        self.lbl_remarque.setStyleSheet(
            f"QLabel{{background:{t['card_bg']};border:1px solid {t['border']};"
            f"border-left:3px solid {t['remarque_border']};border-radius:6px;"
            f"padding:8px 12px;color:{t['text']};font-size:11px;}}"
        )
        self.waterfall.apply_theme()
        if self._client:
            ms = matiere_style(self._matiere)
            self.lbl_matiere_chip.setStyleSheet(
                f"background:{ms['soft_bg']};color:{ms['soft_fg']};border-radius:9px;"
                f"padding:3px 9px;letter-spacing:.5px;"
            )
            self._refresh_chart()

    def load_client(self, client: str, matiere: str):
        """
        Charge un nouveau client dans le panneau :
        - cache le placeholder et affiche le contenu
        - met à jour l'en-tête et la cascade
        - relance le calcul si un PL21 est déjà saisi
        """
        self._client   = client
        self._matiere  = matiere
        self._last_res = None   # évite qu'une cascade stale (autre client) ne s'affiche
        self._placeholder.hide()
        self._content.show()
        self.lbl_heading.setText(client)
        ms = matiere_style(matiere)
        self.lbl_matiere_chip.setText(matiere)
        self.lbl_matiere_chip.setStyleSheet(
            f"background:{ms['soft_bg']};color:{ms['soft_fg']};border-radius:9px;"
            f"padding:3px 9px;letter-spacing:.5px;"
        )
        # On cache le résultat jusqu'au calcul (la remise, elle, est déjà connue)
        self.btn_copier.hide()
        self.card_total.set_value("—")
        self.card_final.set_value("—")
        remise = DATA["remises"][matiere][client]
        self.card_remise.set_value(_fmt_remise(abs(remise) * 100))
        self._update_remarque()
        self._refresh_chart()
        # Si l'utilisateur avait déjà saisi un PL21, recalculer immédiatement
        if self._pl21 is not None:
            self._calculer(self._pl21)
        self._fade_content()

    def _fade_content(self):
        """
        Fondu d'apparition (160 ms) du contenu quand on change de client — la
        transition brutale devient fluide sans ralentir l'usage. L'effet
        d'opacité est retiré à la fin : le laisser en place peut dégrader le
        rendu des ombres portées des cartes enfants.
        """
        if self._fade_anim is not None:
            self._fade_anim.stop()
            self._fade_anim.deleteLater()
        eff = QGraphicsOpacityEffect(self._content)
        self._content.setGraphicsEffect(eff)
        anim = QPropertyAnimation(eff, b"opacity", self)
        anim.setDuration(160)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.finished.connect(lambda: self._content.setGraphicsEffect(None))
        anim.start()
        self._fade_anim = anim

    def _refresh_chart(self, animate: bool = False):
        """
        Reconstruit les étapes de la cascade (WaterfallChart) pour le client courant
        et les lui transmet. Appelé à chaque changement de client, après une
        modification des taux/de la remise, ou à chaque calcul.
        animate=True (après un calcul) : les colonnes poussent une à une.

        Sans PL21 saisi, la cascade est indexée sur une base 100 : la forme
        (proportions des hausses, poids de la remise) reste correcte, seuls
        les libellés de valeur affichent « — » en attendant un calcul.
        """
        if not self._client:
            return
        augments = DATA["augmentations"][self._matiere][self._client]
        annees   = sorted(augments.keys())
        remise   = DATA["remises"][self._matiere][self._client]
        steps: list[dict] = []

        if self._pl21 is not None and self._last_res is not None:
            res  = self._last_res
            prev = self._pl21
            steps.append({"kind": "start", "name": "PL21", "value": prev,
                          "value_text": f"{prev:.2f} €"})
            for etape in res["etapes"]:
                steps.append({
                    "kind": "increase", "name": etape["annee"],
                    "prev_value": prev, "value": etape["prix_apres"],
                    "delta_text": f"+{etape['taux'] * 100:.1f}%",
                    "value_text": f"{etape['prix_apres']:.2f} €",
                })
                prev = etape["prix_apres"]
            steps.append({
                "kind": "final", "name": "Prix final",
                "prev_value": prev, "value": res["prix_final"],
                "delta_text": f"-{res['remise_pct']}%",
                "value_text": f"{res['prix_final']} €",
            })
        else:
            prev = 100.0
            steps.append({"kind": "start", "name": "PL21", "value": prev, "value_text": "—"})
            for annee in annees:
                taux = augments[annee]
                val  = prev * (1 + taux)
                steps.append({
                    "kind": "increase", "name": annee, "prev_value": prev, "value": val,
                    "delta_text": f"+{taux * 100:.1f}%", "value_text": "—",
                })
                prev = val
            steps.append({
                "kind": "final", "name": "Prix final", "prev_value": prev,
                "value": max(prev * (1 + remise), prev * 0.01),
                "delta_text": f"{remise * 100:.1f}%", "value_text": "—",
            })

        self.waterfall.set_steps(steps, self._matiere, animate)

    def _calculer(self, pl21: float):
        """Appelle calculer_prix(), met à jour les cartes de synthèse et la cascade."""
        if not self._client:
            return
        res = calculer_prix(pl21, self._matiere, self._client)
        self._last_res = res
        augmentation_pct = (res["prix_avant_remise"] / pl21 - 1) * 100
        # Effet compteur sur les valeurs calculées ; la remise (déjà connue) reste fixe
        self.card_total.animate_value(augmentation_pct, lambda v: f"+{v:.1f}%")
        self.card_remise.set_value(f"-{res['remise_pct']}%")
        self.card_final.animate_value(float(res["prix_final"]), lambda v: f"{v:.0f} €")
        self.btn_copier.show()
        self._refresh_chart(animate=True)

    def update_pl21(self, pl21: float | None):
        """
        Point d'entrée appelé par MainWindow quand le PL21 change.
        Si pl21 est None (réinitialisation), efface les prix calculés.
        """
        self._pl21 = pl21
        if pl21 is not None:
            self._calculer(pl21)
        else:
            self._reset_prices()

    def _reset_prices(self):
        """Remet les cartes dépendantes du PL21 à « — » et la cascade à sa forme indexée."""
        self.btn_copier.hide()
        self.card_total.set_value("—")
        self.card_final.set_value("—")
        self._last_res = None
        self._refresh_chart()

    def _ouvrir_remise_dialog(self):
        """
        Ouvre une mini boîte de dialogue pour modifier la remise du client courant.
        Aucun mot de passe n'est requis ici (usage quotidien fréquent).
        """
        if not self._client:
            return
        t = theme()
        remise_pct = abs(DATA["remises"][self._matiere][self._client]) * 100
        dlg = QDialog(self)
        dlg.setWindowTitle("Taux de remise")
        dlg.setFixedWidth(280)
        dlg.setModal(True)
        dlg.setStyleSheet(f"QDialog{{background:{t['card_bg']};}}QLabel{{color:{t['text']};}}")
        v = QVBoxLayout(dlg)
        v.setContentsMargins(16, 16, 16, 16)
        v.setSpacing(12)
        lbl = QLabel(f"Modifier la remise — {self._client} ({self._matiere})")
        lbl.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        lbl.setStyleSheet(f"color:{t['accent']};")
        v.addWidget(lbl)
        spin = QDoubleSpinBox()
        spin.setRange(0, 99)
        spin.setDecimals(1)
        spin.setSuffix(" %")
        spin.setValue(remise_pct)
        spin.setFixedHeight(32)
        spin.setFont(QFont("Segoe UI", 11))
        v.addWidget(spin)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Appliquer")
        btns.button(QDialogButtonBox.StandardButton.Ok).setStyleSheet(
            f"background:{t['accent']};color:white;padding:5px 14px;border-radius:4px;border:none;"
        )
        btns.accepted.connect(lambda: self._appliquer_remise(spin.value(), dlg))
        btns.rejected.connect(dlg.reject)
        v.addWidget(btns)
        dlg.exec()
        dlg.deleteLater()

    def _appliquer_remise(self, valeur_pct: float, dlg=None):
        """
        Sauvegarde le nouveau taux de remise, met à jour l'affichage
        et relance le calcul si un PL21 est actif.
        """
        DATA["remises"][self._matiere][self._client] = -round(valeur_pct / 100, 4)
        # log_audit() persiste DATA : pas besoin d'un save_data() séparé ici.
        log_audit("edit_remise", self._matiere, self._client, _fmt_remise(valeur_pct))
        self.card_remise.set_value(_fmt_remise(valeur_pct))
        if dlg:
            dlg.accept()
        if self._pl21 is not None:
            self._calculer(self._pl21)
        else:
            self._refresh_chart()

    def _copier_prix(self):
        """
        Copie le prix final (entier) dans le presse-papiers système.
        La confirmation s'affiche dans un toast éphémère (voir MainWindow.show_toast).
        """
        if not self._last_res:
            return
        QApplication.clipboard().setText(str(self._last_res["prix_final"]))
        self.main_window.show_toast(f"✓ Prix copié : {self._last_res['prix_final']} €")

    def _export_pdf_client(self):
        """
        Génère un PDF pour le client actuellement sélectionné.

        Si un PL21 est calculé : le PDF inclut le tableau complet avec les prix
        intermédiaires, la remise et le prix final mis en évidence.
        Sans PL21 : le PDF montre les taux et la remise sans les prix (colonnes « — »).

        Le fichier est généré via QTextDocument (HTML → PDF) avec QPrinter.
        """
        if not self._client:
            return

        # Nom de fichier par défaut : le nom du client est du texte libre (voir
        # AddClientDialog._valider), donc pas garanti compatible avec les noms
        # de fichiers Windows — on retire tous les caractères interdits, pas
        # seulement l'espace et le slash.
        nom_fichier = _re.sub(
            r'[\\/:*?"<>|]', '_',
            f"chiffrage_{self._client}_{self._matiere}_{datetime.date.today()}.pdf"
            .replace(" ", "_")
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter PDF client", nom_fichier, "PDF (*.pdf)"
        )
        if not path:
            return

        # Construction du corps HTML du PDF (palette fixe, indépendante du thème écran)
        pp = print_palette(self._matiere)
        if self._pl21 is not None:
            # Cas avec PL21 calculé : tableau complet avec prix
            res       = calculer_prix(self._pl21, self._matiere, self._client)
            pl21_txt  = f"{self._pl21:.2f} €"
            lignes_html = ""
            for i, etape in enumerate(res["etapes"]):
                bg = pp["stripe"] if i % 2 else "#ffffff"
                lignes_html += (
                    f"<tr style='background:{bg};'>"
                    f"<td style='padding:7px 12px;'>{etape['annee']}</td>"
                    f"<td style='padding:7px 12px;text-align:center;'>+{etape['taux']*100:.1f}%</td>"
                    f"<td style='padding:7px 12px;text-align:right;color:{pp['accent']};'>{etape['prix_apres']:.2f} €</td>"
                    f"</tr>"
                )
            lignes_html += (
                f"<tr style='background:{pp['success_bg']};font-weight:bold;'>"
                f"<td style='padding:7px 12px;'>Remise</td>"
                f"<td style='padding:7px 12px;text-align:center;color:{pp['danger_fg']};'>-{res['remise_pct']}%</td>"
                f"<td style='padding:7px 12px;text-align:right;color:{pp['success_fg']};'>{res['prix_final']} €</td>"
                f"</tr>"
            )
            prix_final_html = (
                f"<p style='margin-top:18px;background:{pp['success_bg']};border:2px solid {pp['success_border']};"
                f"border-radius:6px;padding:12px;font-size:14pt;font-weight:bold;color:{pp['success_fg']};"
                f"text-align:center;'>Prix unitaire final : {res['prix_final']} €</p>"
            )
        else:
            # Cas sans PL21 : fiche de taux uniquement
            pl21_txt    = "—"
            augments    = DATA["augmentations"][self._matiere][self._client]
            remise      = abs(DATA["remises"][self._matiere][self._client]) * 100
            lignes_html = ""
            for i, (annee, taux) in enumerate(sorted(augments.items())):
                bg = pp["stripe"] if i % 2 else "#ffffff"
                lignes_html += (
                    f"<tr style='background:{bg};'>"
                    f"<td style='padding:7px 12px;'>{annee}</td>"
                    f"<td style='padding:7px 12px;text-align:center;'>+{taux*100:.1f}%</td>"
                    f"<td style='padding:7px 12px;text-align:right;color:{pp['ink_faint']};'>—</td>"
                    f"</tr>"
                )
            lignes_html += (
                f"<tr style='background:{pp['success_bg']};font-weight:bold;'>"
                f"<td style='padding:7px 12px;'>Remise</td>"
                f"<td style='padding:7px 12px;text-align:center;color:{pp['danger_fg']};'>{_fmt_remise(remise)}</td>"
                f"<td style='padding:7px 12px;text-align:right;color:{pp['ink_faint']};'>—</td>"
                f"</tr>"
            )
            prix_final_html = ""

        # Bloc remarque (absent si pas de remarque pour ce client). Échappé (html.escape)
        # comme le nom du client et de la matière ci-dessous : ce sont des textes libres
        # saisis par un admin via EditClientDialog/AddClientDialog, pas du HTML de confiance
        # — un caractère "&", "<" ou ">" y casserait sinon le rendu du PDF généré.
        remarque = DATA["remarques"].get(self._client, "")
        remarque_html = (
            f"<p style='background:{pp['remarque_bg']};border:1px solid {pp['line']};"
            f"border-left:3px solid {pp['remarque_border']};border-radius:5px;"
            f"padding:10px;color:{pp['ink']};font-size:9pt;margin-top:14px;'>"
            f"{_html.escape(remarque)}</p>"
        ) if remarque else ""
        client_html  = _html.escape(self._client)
        matiere_html = _html.escape(self._matiere)

        html = f"""
        <html><body style='font-family:Arial;font-size:10pt;color:{pp['ink']};'>
          <table width='100%'><tr>
            <td>
              <h2 style='color:{pp['accent']};margin:0;'>Chiffrage OEM — {client_html}</h2>
              <p style='margin:4px 0 0 0;color:{pp['ink_muted']};'>
                Matière : {matiere_html} &nbsp;|&nbsp;
                Date : {datetime.date.today().strftime('%d/%m/%Y')} &nbsp;|&nbsp;
                PL21 : {pl21_txt}
              </p>
            </td>
          </tr></table>
          <hr style='border:1px solid {pp['line']};margin:12px 0;'>
          <table width='100%' border='1' cellspacing='0'
                 style='border-collapse:collapse;border-color:{pp['line']};'>
            <tr style='background:{pp['accent']};color:white;'>
              <th style='padding:8px 12px;text-align:left;'>Année</th>
              <th style='padding:8px 12px;text-align:center;'>Taux appliqué</th>
              <th style='padding:8px 12px;text-align:right;'>Prix après (€)</th>
            </tr>
            {lignes_html}
          </table>
          {prix_final_html}
          {remarque_html}
        </body></html>
        """

        if export_html_to_pdf(self, html, path):
            QMessageBox.information(self, "Export PDF", f"PDF exporté :\n{path}")

    def _ouvrir_edition(self):
        """
        Ouvre EditClientDialog après vérification du mot de passe admin.
        Reconstruit la cascade et relance le calcul si nécessaire après validation.
        """
        if not self._client:
            return
        if not verifier_admin(self):
            return
        dlg = EditClientDialog(self._matiere, self._client, self)
        accepted = dlg.exec() == QDialog.DialogCode.Accepted
        dlg.deleteLater()
        if accepted:
            self._refresh_chart()
            if self._pl21 is not None:
                self._calculer(self._pl21)
            self._update_remarque()
            QMessageBox.information(self, "Sauvegardé", "Modifications enregistrées.")

    def _update_remarque(self):
        """Affiche ou cache le bloc de remarque selon qu'une remarque existe pour ce client."""
        rem = DATA["remarques"].get(self._client, "") if self._client else ""
        if rem:
            accent = theme()["remarque_border"]
            self.lbl_remarque.setText(
                f"<span style='color:{accent};font-weight:700;letter-spacing:1px;"
                f"font-size:9px;'>NOTE</span><br>{rem}"
            )
            self.lbl_remarque.show()
        else:
            self.lbl_remarque.hide()

    def clear_client(self):
        """
        Réinitialise le panneau à son état vide (aucun client sélectionné).
        Appelé quand une matière ne contient plus aucun client (dernier client
        supprimé) : évite de garder affiché un client qui n'existe plus dans
        DATA, ce qui provoquerait un KeyError au clic sur PDF/Modifier/remise.
        """
        self._client   = None
        self._matiere  = None
        self._pl21     = None
        self._last_res = None
        self.waterfall.set_steps([], None)
        self._content.hide()
        self._placeholder.show()


# ─────────────────────────────────────────────────────────────────────────────
#  FENÊTRE PRINCIPALE
# ─────────────────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    """
    Fenêtre principale de l'application.

    Layout vertical :
      1. Header (titre + logo)
      2. Barre matière + PL21 (sélecteur PTFE/PEEK, saisie, Calculer, Réinitialiser, filtre)
      3. QSplitter horizontal
           gauche : sidebar scrollable avec ClientListItem
           droite : ClientDetailPanel

    Flux principal :
      • L'utilisateur clique sur une matière → _switch_matiere() recharge la sidebar
      • L'utilisateur saisit un PL21 et clique Calculer → _calculer_tout() :
            - calcule le prix de TOUS les clients visibles
            - met à jour leur ClientListItem avec le prix obtenu
            - transmet le PL21 au ClientDetailPanel
      • L'utilisateur clique sur un client dans la sidebar → _select_client() :
            - met en surbrillance l'item sélectionné
            - charge ce client dans le ClientDetailPanel
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Procédure Chiffrage OEM — Anneaux / Segmentation")
        self.setMinimumSize(800, 560)

        # Matière affichée au démarrage = première clé de DATA["augmentations"]
        self.matiere_courante = list(DATA["augmentations"].keys())[0]
        self._matiere_buttons: dict              = {}   # { "PTFE": QPushButton, ... }
        self._m_layout        = None                    # QHBoxLayout de la barre matières
        self._client_items: dict[str, ClientListItem] = {}  # { "Atlas": ClientListItem, ... }
        self._selected_client: str | None        = None
        self._pl21:            float | None      = None  # PL21 courant (None = non calculé)

        self._build_ui()
        self._build_menu()
        self._toast = Toast(self)
        self._setup_shortcuts()
        if CORRUPT_DATA_BACKUP is not None:
            # Différé après l'affichage de la fenêtre (QTimer.singleShot(0, …))
            # pour ne pas bloquer une QMessageBox modale avant que l'UI existe.
            QTimer.singleShot(0, self._warn_corrupt_data)

    def _warn_corrupt_data(self):
        """Avertit une fois que le fichier de données était illisible (voir load_data)."""
        QMessageBox.warning(
            self, "Données réinitialisées",
            "Le fichier de données existant était illisible ou dans un format "
            "inattendu (JSON corrompu ou structure invalide).\n\n"
            "L'application est repartie sur des données par défaut. L'ancien "
            f"fichier a été conservé ici pour investigation :\n{CORRUPT_DATA_BACKUP}"
        )

    def _setup_shortcuts(self):
        """
        Raccourcis clavier de confort :
          Ctrl+F : focus sur le filtre clients
          ↑ / ↓  : navigation parmi les clients visibles de la sidebar
        """
        sc_find = QShortcut(QKeySequence.StandardKey.Find, self)
        sc_find.activated.connect(self._focus_filter)
        sc_up = QShortcut(QKeySequence(Qt.Key.Key_Up), self)
        sc_up.activated.connect(lambda: self._nav_client(-1))
        sc_down = QShortcut(QKeySequence(Qt.Key.Key_Down), self)
        sc_down.activated.connect(lambda: self._nav_client(1))

    def _focus_filter(self):
        self.search_input.setFocus()
        self.search_input.selectAll()

    def _nav_client(self, delta: int):
        """Sélectionne le client précédent/suivant (↑/↓) parmi les items visibles."""
        visibles = [c for c, it in self._client_items.items() if it.isVisible()]
        if not visibles:
            return
        if self._selected_client in visibles:
            idx = visibles.index(self._selected_client) + delta
            idx = max(0, min(len(visibles) - 1, idx))
        else:
            idx = 0
        client = visibles[idx]
        if client == self._selected_client:
            return
        self._select_client(client)
        self._clients_scroll.ensureWidgetVisible(self._client_items[client], 0, 30)

    def show_toast(self, text: str):
        """Affiche une notification éphémère en bas de la fenêtre."""
        self._toast.show_message(text)

    # ── Construction de l'interface ───────────────────────────────────────────

    def _build_ui(self):
        """Construit toute l'interface graphique de la fenêtre principale."""
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)

        # ── Bandeau header (dégradé + sous-titre + bouton thème) ───────────────
        self._header = QFrame()
        self._header.setObjectName("header")
        self._header.setMinimumHeight(72)
        self._header.setMaximumHeight(84)
        hl = QHBoxLayout(self._header)
        hl.setContentsMargins(24, 10, 24, 10)
        hl.setSpacing(16)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self.lbl_title = QLabel("Procédure Chiffrage OEM")
        self.lbl_title.setFont(font_display(16))
        self.lbl_subtitle = QLabel("Anneaux & segments  ·  CPI by Howden")
        self.lbl_subtitle.setFont(QFont("Segoe UI", 9))
        self.lbl_subtitle.setToolTip(
            "Créé par Dylan Carlier, imaginé par Nicolas Richet et Robin Demeure."
        )
        title_col.addWidget(self.lbl_title)
        title_col.addWidget(self.lbl_subtitle)
        hl.addLayout(title_col)
        hl.addStretch()

        # Bouton de bascule clair/sombre
        self.btn_theme = QPushButton()
        self.btn_theme.setFixedSize(32, 32)
        self.btn_theme.setIconSize(QSize(16, 16))
        self.btn_theme.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_theme.clicked.connect(self._toggle_theme)
        hl.addWidget(self.btn_theme)

        # Logo : cherché dans le même dossier que le script sous le nom logo_cpi.png
        logo_path = os.path.join(str(_app_dir()), "logo_cpi.png")
        logo_lbl  = QLabel()
        logo_lbl.setFixedSize(130, 52)
        logo_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_lbl.setStyleSheet("background:white;border-radius:6px;padding:4px;")
        pix = QPixmap(logo_path)
        if not pix.isNull():
            pix = pix.scaledToHeight(44, Qt.TransformationMode.SmoothTransformation)
            logo_lbl.setPixmap(pix)
            logo_lbl.setFixedSize(pix.width() + 12, 52)
        else:
            # Fallback texte si le fichier logo n'est pas trouvé
            logo_lbl.setText("CPI by Howden")
            logo_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
            logo_lbl.setStyleSheet(
                "color:#3e6c8e;background:white;border-radius:6px;padding:4px;"
            )
        hl.addWidget(logo_lbl)
        root.addWidget(self._header)

        # ── Barre matière + PL21 + actions ──────────────────────────────────────
        # Une seule ligne (matière, PL21, Calculer, Réinitialiser, filtre) : le
        # choix de matière et la saisie du prix sont la même décision, pas deux
        # étapes séparées dans deux bandeaux différents.
        self._top_bar = QFrame()
        self._top_bar.setFixedHeight(52)
        tl = QHBoxLayout(self._top_bar)
        tl.setContentsMargins(20, 8, 20, 8)
        tl.setSpacing(10)

        self.lbl_mat = QLabel("Matière :")
        self.lbl_mat.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        tl.addWidget(self.lbl_mat)

        # Groupe segmenté : les boutons matière partagent un même cadre, sans
        # espace entre eux, pour former un vrai sélecteur à onglets (pas des
        # pastilles isolées côte à côte).
        self._matiere_group = QFrame()
        self._matiere_group.setObjectName("matiereGroup")
        self._mg_layout = QHBoxLayout(self._matiere_group)
        self._mg_layout.setContentsMargins(0, 0, 0, 0)
        self._mg_layout.setSpacing(0)
        tl.addWidget(self._matiere_group)

        # Créer un bouton par matière présente dans DATA
        for mat in DATA["augmentations"]:
            self._create_matiere_btn(mat)

        # Repère « PL21 » + valeur regroupés dans un seul cadre façon afficheur
        # d'instrument, au lieu d'un libellé et d'un champ séparés.
        self._pl21_gauge = QFrame()
        self._pl21_gauge.setObjectName("pl21Gauge")
        self._pl21_gauge.setFixedSize(270, 34)
        pg = QHBoxLayout(self._pl21_gauge)
        pg.setContentsMargins(12, 0, 10, 0)
        pg.setSpacing(8)
        self.lbl_pl21 = QLabel("PL21")
        self.lbl_pl21.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        pg.addWidget(self.lbl_pl21)
        # Champ de saisie PL21 — accepte un nombre ou une expression (voir evaluer_expression)
        self.input_pl21 = QLineEdit()
        self.input_pl21.setPlaceholderText("ex: 100  ou  600 + 50  ou  620 + 20%")
        self.input_pl21.setFixedWidth(220)
        self.input_pl21.setFrame(False)
        self.input_pl21.setFont(font_mono(11))
        # Prévisualisation en temps réel de l'expression saisie
        self.input_pl21.textChanged.connect(self._preview_pl21)
        pg.addWidget(self.input_pl21)
        tl.addWidget(self._pl21_gauge)

        # Label de prévisualisation (affiché à droite de l'afficheur)
        self.lbl_preview = QLabel("")
        self.lbl_preview.setFont(font_mono(9))
        self.lbl_preview.hide()
        tl.addWidget(self.lbl_preview)

        self.btn_calc = QPushButton("Calculer")
        self.btn_calc.setFixedHeight(34)
        self.btn_calc.setFixedWidth(100)
        self.btn_calc.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.btn_calc.clicked.connect(self._calculer_tout)
        # Appuyer sur Entrée dans le champ PL21 déclenche aussi le calcul
        self.input_pl21.returnPressed.connect(self._calculer_tout)
        tl.addWidget(self.btn_calc)

        self.btn_reset = QPushButton("Réinitialiser")
        self.btn_reset.setFixedHeight(34)
        self.btn_reset.setFont(QFont("Segoe UI", 10))
        self.btn_reset.clicked.connect(self._reset_tout)
        tl.addWidget(self.btn_reset)

        tl.addStretch()

        # Filtre texte pour la sidebar (masque les clients ne correspondant pas)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Filtrer clients…")
        self.search_input.setFixedWidth(180)
        self.search_input.setFixedHeight(32)
        self.search_input.textChanged.connect(self._filter_clients)
        tl.addWidget(self.search_input)
        root.addWidget(self._top_bar)

        # ── Splitter principal ────────────────────────────────────────────────
        self._splitter = QSplitter(Qt.Orientation.Horizontal)
        self._splitter.setHandleWidth(1)
        splitter = self._splitter

        # Panneau gauche : sidebar des clients
        self._sidebar = QWidget()
        sidebar = self._sidebar
        sidebar.setMinimumWidth(190)
        sidebar.setMaximumWidth(280)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setSpacing(0)
        sidebar_layout.setContentsMargins(0, 8, 0, 8)

        self.lbl_clients = QLabel("CLIENTS")
        self.lbl_clients.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        sidebar_layout.addWidget(self.lbl_clients)

        # Affiché à la place de la liste quand le filtre ne retient aucun client
        self.lbl_no_result = QLabel("Aucun client ne correspond")
        _f_no_result = QFont("Segoe UI", 9)
        _f_no_result.setItalic(True)
        self.lbl_no_result.setFont(_f_no_result)
        self.lbl_no_result.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_no_result.hide()
        sidebar_layout.addWidget(self.lbl_no_result)

        # Zone scrollable contenant les ClientListItem
        self._clients_scroll = QScrollArea()
        self._clients_scroll.setWidgetResizable(True)
        self._clients_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self._clients_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._clients_container = QWidget()
        self._clients_vlayout   = QVBoxLayout(self._clients_container)
        self._clients_vlayout.setSpacing(0)
        self._clients_vlayout.setContentsMargins(0, 0, 0, 0)
        self._clients_scroll.setWidget(self._clients_container)
        sidebar_layout.addWidget(self._clients_scroll)
        splitter.addWidget(sidebar)

        # Panneau droit : détail du client sélectionné
        self._detail_panel = ClientDetailPanel(self)
        splitter.addWidget(self._detail_panel)

        # Taille initiale : sidebar 220 px, détail prend le reste
        splitter.setSizes([220, 700])
        splitter.setStretchFactor(0, 0)   # sidebar : largeur fixe
        splitter.setStretchFactor(1, 1)   # détail : prend tout l'espace restant
        root.addWidget(splitter)

        # Barre de statut : contexte courant (matière, client, PL21, nb clients affichés)
        self.setStatusBar(QStatusBar())

        # Remplissage initial de la sidebar, puis peinture des couleurs du thème actif
        self._populate_clients(self.matiere_courante)
        self._apply_theme()

    # ── Thème clair/sombre ────────────────────────────────────────────────────

    def _toggle_theme(self):
        """
        Bascule entre thème clair et sombre avec un fondu croisé (~260 ms) :
        l'ancien rendu de la fenêtre est capturé en image, affiché en overlay
        au-dessus du nouveau thème, puis s'efface — le changement de thème
        devient une transition douce au lieu d'un clignotement.
        """
        old_pix = self.grab()
        global CURRENT_THEME
        CURRENT_THEME = "dark" if CURRENT_THEME == "light" else "light"
        self._apply_theme()

        overlay = QLabel(self)
        overlay.setPixmap(old_pix)
        overlay.resize(self.size())
        overlay.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        overlay.show()
        overlay.raise_()
        eff = QGraphicsOpacityEffect(overlay)
        overlay.setGraphicsEffect(eff)
        anim = QPropertyAnimation(eff, b"opacity", overlay)
        anim.setDuration(260)
        anim.setStartValue(1.0)
        anim.setEndValue(0.0)
        anim.setEasingCurve(QEasingCurve.Type.InOutQuad)
        anim.finished.connect(overlay.deleteLater)
        anim.start()

    def _apply_theme(self):
        """
        Réapplique les couleurs du thème actif à tous les éléments de chrome
        (header, barres, sidebar, statusbar) ainsi qu'aux widgets enfants
        (ClientListItem, ClientDetailPanel) qui savent se repeindre eux-mêmes.
        """
        t = theme()
        app = QApplication.instance()
        palette = app.palette()
        palette.setColor(QPalette.ColorRole.Window,          QColor(t["window_bg"]))
        palette.setColor(QPalette.ColorRole.WindowText,      QColor(t["text"]))
        palette.setColor(QPalette.ColorRole.Base,            QColor(t["card_bg"]))
        palette.setColor(QPalette.ColorRole.AlternateBase,   QColor(t["sidebar_bg"]))
        palette.setColor(QPalette.ColorRole.Text,            QColor(t["text"]))
        palette.setColor(QPalette.ColorRole.Button,          QColor(t["btn_secondary_bg"]))
        palette.setColor(QPalette.ColorRole.ButtonText,      QColor(t["btn_secondary_fg"]))
        palette.setColor(QPalette.ColorRole.Highlight,       QColor(t["accent"]))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        app.setPalette(palette)

        # Header — bandeau plat « papier », plus de dégradé bleu marine : l'identité
        # de la fenêtre vient désormais de la couleur de matière (voir plus bas),
        # pas d'un bandeau de marque façon SaaS.
        self._header.setStyleSheet(
            f"QFrame#header{{background:{t['card_bg']};border-bottom:1px solid {t['border']};}}"
        )
        self.lbl_title.setStyleSheet(f"color:{t['text']};")
        self.lbl_subtitle.setStyleSheet(f"color:{t['text_secondary']};")
        self.btn_theme.setIcon(make_theme_icon("sun" if CURRENT_THEME == "dark" else "moon", t["text"]))
        self.btn_theme.setStyleSheet(
            f"QPushButton{{background:transparent;border:1px solid {t['border']};border-radius:16px;}}"
            f"QPushButton:hover{{background:{t['btn_secondary_hover']};}}"
        )

        # Barre matière + PL21 (une seule ligne, pas de ligne de séparation en bas :
        # le contraste de fond avec la sidebar/le contenu suffit à la distinguer).
        self._top_bar.setStyleSheet(f"background:{t['card_bg']};border:none;")
        self.lbl_mat.setStyleSheet(f"color:{t['text_secondary']};")
        self._matiere_group.setStyleSheet("QFrame#matiereGroup{background:transparent;border:none;}")
        self._restyle_matiere_group()

        # Afficheur PL21 : un seul cadre net (comme la maquette) qui regroupe le
        # libellé et la valeur — le champ de saisie lui-même n'a aucune bordure
        # propre, y compris quand il a le focus, pour ne pas créer de « boîte
        # dans la boîte ».
        self._pl21_gauge.setStyleSheet(
            f"QFrame#pl21Gauge{{background:{t['card_bg']};border:1px solid {t['input_border']};"
            f"border-radius:8px;}}"
        )
        self.lbl_pl21.setStyleSheet(f"color:{t['text_muted']};letter-spacing:.5px;")
        self.input_pl21.setStyleSheet(
            f"QLineEdit{{border:none;background:transparent;color:{t['text']};padding:0;}}"
            f"QLineEdit:focus{{border:none;outline:none;}}"
        )
        self.lbl_preview.setStyleSheet(f"color:{t['text_secondary']};font-size:9pt;border:none;")
        self._style_calc_button()
        self.btn_reset.setStyleSheet(
            f"QPushButton{{background:{t['btn_secondary_bg']};color:{t['btn_secondary_fg']};"
            f"border-radius:5px;padding:0 10px;}}"
            f"QPushButton:hover{{background:{t['btn_secondary_hover']};}}"
        )
        self.search_input.setStyleSheet(
            f"padding:4px 8px;border:1px solid {t['input_border']};border-radius:4px;"
            f"background:{t['card_bg']};color:{t['text']};"
        )

        # Sidebar
        self._sidebar.setStyleSheet(f"background:{t['sidebar_bg']};")
        self.lbl_clients.setStyleSheet(
            f"color:{t['text_muted']};padding:0 16px 8px 16px;letter-spacing:1px;"
        )
        self.lbl_no_result.setStyleSheet(f"color:{t['text_muted']};padding:12px 8px;")
        self._clients_scroll.setStyleSheet(scrollbar_qss())
        self._splitter.setStyleSheet(f"QSplitter::handle{{background:{t['splitter_handle']};}}")

        # Barre de statut
        self.statusBar().setStyleSheet(
            f"QStatusBar{{background:{t['statusbar_bg']};color:{t['statusbar_fg']};"
            f"padding:2px 12px;font-size:9pt;}}"
        )

        # Widgets enfants qui savent se repeindre eux-mêmes
        for item in self._client_items.values():
            item.apply_theme()
        self._detail_panel.apply_theme()

        # Barre de titre Windows alignée sur le thème (sombre/clair)
        apply_titlebar_theme(self)

    def _build_menu(self):
        """Construit la barre de menus (Fichier + Administration)."""
        mb = self.menuBar()

        # Menu Fichier
        fichier = mb.addMenu("Fichier")
        fichier.addAction(self._action("Exporter CSV…",   self._export_csv))
        fichier.addAction(self._action("Exporter Excel…", self._export_xlsx))
        fichier.addAction(self._action("Exporter PDF…",   self._export_pdf))
        fichier.addSeparator()
        fichier.addAction(self._action(
            "Historique des calculs…", self._afficher_historique
        ))
        fichier.addSeparator()
        fichier.addAction(self._action("Quitter", self.close))

        # Menu Administration (toutes les actions nécessitent le mot de passe)
        admin = mb.addMenu("Administration")
        admin.addAction(self._action("Ajouter un client…",   self._admin_ajouter_client))
        admin.addAction(self._action("Supprimer un client…", self._admin_supprimer_client))
        admin.addSeparator()
        admin.addAction(self._action("Ajouter une matière…", self._admin_ajouter_matiere))
        admin.addSeparator()
        admin.addAction(self._action("Changer le mot de passe…",   self._admin_changer_mdp))
        admin.addAction(self._action(
            "Journal des modifications…", self._afficher_audit_log
        ))

    def _afficher_historique(self):
        dlg = HistoriqueDialog(self)
        dlg.exec()
        dlg.deleteLater()

    def _afficher_audit_log(self):
        dlg = AuditLogDialog(self)
        dlg.exec()
        dlg.deleteLater()

    def _action(self, label: str, slot) -> QAction:
        """
        Crée un QAction avec 'self' comme parent (obligatoire pour éviter que Python
        garbage-collect l'objet avant que l'utilisateur puisse cliquer dessus).
        """
        a = QAction(label, self)
        a.triggered.connect(slot)
        return a

    # ── Gestion des matières ──────────────────────────────────────────────────

    def _create_matiere_btn(self, matiere: str):
        """Crée et ajoute un bouton de matière au groupe segmenté PTFE/PEEK/…"""
        btn = QPushButton(matiere)
        btn.setFixedHeight(30)
        btn.setMinimumWidth(76)
        btn.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        btn.clicked.connect(lambda _, m=matiere: self._switch_matiere(m))
        self._matiere_buttons[matiere] = btn
        self._mg_layout.addWidget(btn)
        self._restyle_matiere_group()

    def _restyle_matiere_group(self):
        """
        Réapplique le style de tous les boutons matière, en recalculant à chaque
        fois quels coins arrondir (bord extérieur du groupe uniquement) — les
        boutons doivent former un seul sélecteur segmenté, pas des pastilles
        isolées, y compris après l'ajout d'une matière par l'admin.
        """
        mats = list(self._matiere_buttons.keys())
        for i, mat in enumerate(mats):
            if len(mats) == 1:
                position = "only"
            elif i == 0:
                position = "first"
            elif i == len(mats) - 1:
                position = "last"
            else:
                position = "middle"
            self._set_mat_style(self._matiere_buttons[mat], mat, mat == self.matiere_courante, position)

    def _set_mat_style(self, btn, matiere: str, active: bool, position: str = "only"):
        """
        Applique le style d'un bouton matière selon le thème et sa position dans
        le groupe segmenté (seuls les coins extérieurs du groupe sont arrondis).
        Actif : rempli avec la couleur propre à cette matière (PTFE/PEEK/…) plutôt
        qu'un accent générique — la couleur du bouton annonce la matière choisie.
        Un petit disque coloré (icône peinte, pas d'emoji) rappelle cette couleur
        même à l'état inactif, pour reconnaître les matières d'un coup d'œil.
        """
        radius = {
            "only":   "border-radius:6px;",
            "first":  "border-top-left-radius:6px;border-bottom-left-radius:6px;"
                      "border-top-right-radius:0;border-bottom-right-radius:0;",
            "last":   "border-top-right-radius:6px;border-bottom-right-radius:6px;"
                      "border-top-left-radius:0;border-bottom-left-radius:0;",
            "middle": "border-radius:0;",
        }[position]
        ms = matiere_style(matiere)
        btn.setIcon(make_dot_icon(ms["on"] if active else ms["solid"]))
        btn.setIconSize(QSize(8, 8))
        if active:
            btn.setStyleSheet(
                f"QPushButton{{background:{ms['solid']};color:{ms['on']};{radius}border:none;}}"
            )
        else:
            t = theme()
            btn.setStyleSheet(
                f"QPushButton{{background:{t['btn_secondary_bg']};color:{t['btn_secondary_fg']};"
                f"{radius}border:none;}}"
                f"QPushButton:hover{{background:{t['btn_secondary_hover']};}}"
            )

    def _style_calc_button(self):
        """Colore le bouton Calculer avec la couleur de la matière active."""
        ms = matiere_style(self.matiere_courante)
        self.btn_calc.setStyleSheet(
            f"QPushButton{{background:{ms['solid']};color:{ms['on']};border-radius:5px;}}"
            f"QPushButton:hover{{background:{ms['hover']};}}"
        )

    def _switch_matiere(self, matiere: str):
        """
        Change la matière courante :
        - met à jour le style des boutons (transition de couleur sur Calculer)
        - réinitialise le PL21 et les prix
        - recharge la sidebar avec les clients de la nouvelle matière
        """
        if matiere == self.matiere_courante:
            return
        old_color = QColor(matiere_style(self.matiere_courante)["solid"])
        self.matiere_courante = matiere
        self._restyle_matiere_group()
        self._animate_calc_color(old_color)
        self._reset_tout()
        self._populate_clients(matiere)

    def _animate_calc_color(self, old_color: QColor):
        """
        Fait glisser la couleur du bouton Calculer de l'ancienne matière vers la
        nouvelle (~240 ms) — l'interface « change de température » en douceur au
        lieu de sauter d'un accent à l'autre.
        """
        ms = matiere_style(self.matiere_courante)
        if getattr(self, "_calc_anim", None) is not None:
            self._calc_anim.stop()
            self._calc_anim.deleteLater()
        anim = QVariantAnimation(self)
        anim.setStartValue(old_color)
        anim.setEndValue(QColor(ms["solid"]))
        anim.setDuration(240)
        on = ms["on"]
        anim.valueChanged.connect(lambda c: self.btn_calc.setStyleSheet(
            f"QPushButton{{background:{c.name()};color:{on};border-radius:5px;}}"
        ))
        # À la fin, on repasse par le style normal (qui inclut l'état :hover)
        anim.finished.connect(self._style_calc_button)
        anim.start()
        self._calc_anim = anim

    # ── Gestion de la sidebar ─────────────────────────────────────────────────

    def _populate_clients(self, matiere: str):
        """
        Repeuple la sidebar avec les clients de la matière donnée.
        Sélectionne automatiquement le premier client.

        Un client est toujours ajouté/supprimé dans toutes les matières à la
        fois (voir AddClientDialog/_admin_supprimer_client), donc la liste de
        clients est identique d'une matière à l'autre dans l'immense majorité
        des cas — seule leur couleur/teinte change. Dans ce cas, les widgets
        ClientListItem existants sont réutilisés (juste restylés) plutôt que
        détruits et recréés : évite un churn de dizaines de widgets Qt à
        chaque bascule PTFE/PEEK. Si l'ensemble ou l'ordre diffère (fichier de
        données édité à la main), on retombe sur la reconstruction complète.
        """
        # Toujours resynchroniser matiere_courante ici plutôt que de compter
        # sur l'appelant pour l'avoir déjà fait : _select_client() plus bas
        # s'appuie dessus (via self.matiere_courante), pas sur le paramètre
        # local — sans cette ligne, un appel de _populate_clients() avec une
        # matière différente de l'état courant sélectionnerait un client sous
        # la mauvaise matière et lèverait un KeyError dans load_client().
        self.matiere_courante = matiere
        clients = list(DATA["augmentations"][matiere].keys())
        self._selected_client = None
        self.lbl_no_result.hide()   # tous les clients redeviennent visibles

        if clients and list(self._client_items.keys()) == clients:
            for item in self._client_items.values():
                item.matiere = matiere
                item.clear_prix()
                item.set_active(False)
                item.apply_theme()
        else:
            # Supprimer tous les ClientListItem existants
            while self._clients_vlayout.count():
                item = self._clients_vlayout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            self._client_items.clear()
            for client in clients:
                item = ClientListItem(client, matiere, self._select_client)
                self._clients_vlayout.addWidget(item)
                self._client_items[client] = item
            # Stretch final pour coller les items en haut de la sidebar
            self._clients_vlayout.addStretch()

        # En-tête avec le nombre de clients de cette matière (ex. « 15 CLIENTS »)
        self.lbl_clients.setText(f"{len(self._client_items)} CLIENTS")

        # Sélection automatique du premier client (ou panneau vide s'il n'en reste plus)
        if clients:
            self._select_client(clients[0])
        else:
            self._detail_panel.clear_client()
        self._update_statusbar()

    def _select_client(self, client: str):
        """
        Sélectionne un client dans la sidebar :
        - désactive l'ancien item sélectionné
        - active le nouvel item
        - charge ce client dans le panneau de détail
        """
        if self._selected_client and self._selected_client in self._client_items:
            self._client_items[self._selected_client].set_active(False)
        self._selected_client = client
        if client in self._client_items:
            self._client_items[client].set_active(True)
        self._detail_panel.load_client(client, self.matiere_courante)
        self._update_statusbar()

    def _filter_clients(self, text: str):
        """
        Filtre la sidebar en masquant les clients dont le nom ne contient pas 'text'.
        La recherche est insensible à la casse, aux accents et aux espaces de
        bord (voir _normalize_search). Si aucun client ne correspond, un
        message « Aucun client ne correspond » s'affiche à la place de la liste.
        """
        needle = _normalize_search(text)
        nb_visible = 0
        for client, item in self._client_items.items():
            visible = needle in _normalize_search(client)
            item.setVisible(visible)
            nb_visible += visible
        self.lbl_no_result.setVisible(bool(needle) and nb_visible == 0)
        self._update_statusbar(nb_visible, len(self._client_items))

    # ── Barre de statut ───────────────────────────────────────────────────────

    def _update_statusbar(self, nb_visible: int | None = None, nb_total: int | None = None):
        """
        Met à jour le message de la barre de statut avec le contexte courant.
        nb_visible/nb_total peuvent être passés par l'appelant (ex. _filter_clients,
        qui les a déjà calculés en parcourant les items) pour éviter un second
        parcours complet de la sidebar à chaque frappe dans le filtre.
        """
        if nb_visible is None or nb_total is None:
            nb_visible = sum(1 for it in self._client_items.values() if it.isVisible())
            nb_total   = len(self._client_items)
        pl21_txt   = f"{self._pl21:.2f} €" if self._pl21 is not None else "non calculé"
        client_txt = self._selected_client or "—"
        self.statusBar().showMessage(
            f"Matière : {self.matiere_courante}   •   "
            f"Client : {client_txt}   •   "
            f"PL21 : {pl21_txt}   •   "
            f"{nb_visible}/{nb_total} clients affichés"
        )

    # ── Calcul ────────────────────────────────────────────────────────────────

    def _preview_pl21(self, texte: str):
        """
        Affiche en temps réel le résultat de l'expression saisie dans le champ PL21.
        Masque le label si le champ est vide ou si l'expression est incomplète.
        """
        t = theme()
        if not texte.strip():
            self.lbl_preview.hide()
            return
        try:
            _, label = evaluer_expression(texte)
            self.lbl_preview.setText(f"→ {label} €")
            self.lbl_preview.setStyleSheet(f"color:{t['accent']};font-size:9pt;")
            self.lbl_preview.show()
        except ValueError:
            # L'expression n'est pas encore valide (ex: en cours de saisie)
            try:
                t = texte.replace(",", ".").strip()
                # float() accepte aussi des chiffres Unicode non-ASCII (ex.
                # "１２３", pleine chasse) qu'evaluer_expression (regex ASCII)
                # rejette : sans ce garde-fou, la prévisualisation se cache
                # comme si la saisie était valide, puis "Calculer" échoue.
                if not t.isascii():
                    raise ValueError("chiffres non-ASCII")
                float(t)
                self.lbl_preview.hide()
            except ValueError:
                self.lbl_preview.setText("→ expression en cours…")
                self.lbl_preview.setStyleSheet(f"color:{t['text_muted']};font-size:9pt;")
                self.lbl_preview.show()

    def _calculer_tout(self):
        """
        Déclenché par le bouton « Calculer » ou la touche Entrée.
        - Évalue l'expression PL21 saisie
        - Calcule le prix de tous les clients visibles dans la sidebar
        - Met à jour leurs ClientListItem avec le prix obtenu
        - Transmet le PL21 au panneau de détail
        - Enregistre chaque calcul dans l'historique
        """
        text = self.input_pl21.text().strip()
        if not text:
            QMessageBox.warning(self, "Saisie manquante", "Entrez un prix PL21.")
            return
        try:
            pl21, _ = evaluer_expression(text)
        except ValueError as e:
            QMessageBox.warning(self, "Expression invalide", str(e))
            return

        self._pl21 = pl21
        calcules = False
        for client, item in self._client_items.items():
            if not item.isVisible():
                continue  # ignorer les clients filtrés par la recherche
            res = calculer_prix(pl21, self.matiere_courante, client)
            item.set_prix(res["prix_final"])
            # persist=False : une seule écriture disque après la boucle plutôt
            # qu'une par client (jusqu'à 15 sauvegardes synchrones par clic).
            log_calcul(pl21, self.matiere_courante, client, res["prix_final"], persist=False)
            calcules = True
        if calcules:
            save_data(DATA)
        self._detail_panel.update_pl21(pl21)
        self._update_statusbar()

    def _reset_tout(self):
        """
        Réinitialise le PL21, efface le champ de saisie,
        remet « — » dans tous les items de la sidebar et vide le panneau de détail.
        """
        self._pl21 = None
        self.input_pl21.clear()
        self.lbl_preview.hide()
        for item in self._client_items.values():
            item.clear_prix()
        self._detail_panel.update_pl21(None)
        self._update_statusbar()

    # ── Exports globaux (toute la matière) ───────────────────────────────────

    def _rows_export(self):
        """
        Retourne (liste_clients, liste_années) pour la matière courante.
        La liste des années est l'union de toutes les années de tous les clients,
        triée alphabétiquement (ordre chronologique si format AAAA).
        """
        clients   = list(DATA["augmentations"][self.matiere_courante].keys())
        all_years = sorted({
            y
            for t in DATA["augmentations"][self.matiere_courante].values()
            for y in t
        })
        return clients, all_years

    def _export_csv(self):
        """
        Exporte un fichier CSV (séparateur « ; ») avec tous les clients de la matière courante.
        Si un PL21 est calculé, les colonnes prix avant/après remise sont incluses.
        Encodage UTF-8 avec BOM pour compatibilité Excel.
        """
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter CSV",
            # Le nom de matière est du texte libre (voir _admin_ajouter_matiere) :
            # on retire les caractères interdits dans un nom de fichier Windows.
            _re.sub(r'[\\/:*?"<>|]', '_', f"chiffrage_{self.matiere_courante}_{datetime.date.today()}.csv"),
            "CSV (*.csv)"
        )
        if not path:
            return
        clients, all_years = self._rows_export()
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                # Ce fichier agrège les remises négociées de TOUS les clients
                # de la matière (contrairement à l'export PDF « client »).
                w.writerow(["USAGE INTERNE — NE PAS DIFFUSER"])
                w.writerow(
                    ["Client"] +
                    [f"Taux {y}" for y in all_years] +
                    ["Remise", "Prix av. remise (€)", "Prix final (€)"]
                )
                for client in clients:
                    taux_c = DATA["augmentations"][self.matiere_courante][client]
                    remise = DATA["remises"][self.matiere_courante][client]
                    row = (
                        [csv_safe(client)] +
                        [f"+{taux_c.get(y, 0) * 100:.1f}%" for y in all_years] +
                        [_fmt_remise(abs(remise) * 100)]
                    )
                    if self._pl21:
                        res = calculer_prix(self._pl21, self.matiere_courante, client)
                        row += [f"{res['prix_avant_remise']:.2f}", str(res["prix_final"])]
                    else:
                        row += ["", ""]
                    w.writerow(row)
        except OSError as e:
            # Cause la plus fréquente : le fichier cible est déjà ouvert dans Excel.
            QMessageBox.warning(
                self, "Export impossible",
                f"Le fichier n'a pas pu être écrit (peut-être déjà ouvert ailleurs) :\n{e}"
            )
            return
        QMessageBox.information(self, "Export CSV", f"Exporté :\n{path}")

    def _export_xlsx(self):
        """
        Exporte un fichier Excel (.xlsx) avec mise en forme (en-têtes dans la couleur
        de la matière, prix finaux en vert). Nécessite openpyxl (pip install openpyxl).
        """
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Module manquant", "pip install openpyxl")
            return
        import openpyxl
        from openpyxl.styles import Font as XLFont, PatternFill, Alignment as XLAlign
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter Excel",
            _re.sub(r'[\\/:*?"<>|]', '_', f"chiffrage_{self.matiere_courante}_{datetime.date.today()}.xlsx"),
            "Excel (*.xlsx)"
        )
        if not path:
            return
        clients, all_years = self._rows_export()
        pp = print_palette(self.matiere_courante)
        wb = openpyxl.Workbook()
        ws = wb.active
        # openpyxl rejette les titres de feuille de plus de 31 caractères ou
        # contenant \ / ? * [ ] — or le nom de matière est saisi librement par
        # un admin (voir _admin_ajouter_matiere), donc non garanti conforme.
        titre = _re.sub(r'[\\/?*\[\]]', '_', f"Chiffrage {self.matiere_courante}")[:31]
        ws.title = titre
        fill_h  = PatternFill("solid", fgColor=pp["accent"].lstrip("#").upper())
        font_h  = XLFont(bold=True, color="FFFFFF")
        font_g  = XLFont(bold=True, color=pp["success_fg"].lstrip("#").upper())
        center  = XLAlign(horizontal="center", vertical="center")
        headers = (
            ["Client"] +
            [f"Taux {y}" for y in all_years] +
            ["Remise", "Prix av. remise (€)", "Prix final (€)"]
        )
        # Ce fichier agrège les remises négociées de TOUS les clients de la
        # matière (contrairement à l'export PDF « client » qui n'expose que le
        # sien) : un simple rappel visuel réduit le risque qu'il soit transmis
        # par erreur en dehors de l'entreprise.
        ws.append(["USAGE INTERNE — NE PAS DIFFUSER"])
        ws["A1"].font = XLFont(italic=True, color="808080")
        ws.append(headers)
        for cell in ws[2]:
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = center
        for i, client in enumerate(clients, 3):
            taux_c = DATA["augmentations"][self.matiere_courante][client]
            remise = DATA["remises"][self.matiere_courante][client]
            row = (
                [csv_safe(client)] +
                [f"+{taux_c.get(y, 0) * 100:.1f}%" for y in all_years] +
                [_fmt_remise(abs(remise) * 100)]
            )
            if self._pl21:
                res = calculer_prix(self._pl21, self.matiere_courante, client)
                row += [round(res["prix_avant_remise"], 2), res["prix_final"]]
            else:
                row += ["", ""]
            ws.append(row)
            if self._pl21:
                ws.cell(i, len(headers)).font      = font_g
                ws.cell(i, len(headers)).alignment = center
        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 30
            )
        try:
            wb.save(path)
        except OSError as e:
            # Cause la plus fréquente : le fichier cible est déjà ouvert dans Excel.
            QMessageBox.warning(
                self, "Export impossible",
                f"Le fichier n'a pas pu être écrit (peut-être déjà ouvert ailleurs) :\n{e}"
            )
            return
        QMessageBox.information(self, "Export Excel", f"Exporté :\n{path}")

    def _export_pdf(self):
        """
        Exporte un tableau récapitulatif PDF de tous les clients de la matière courante.
        Utilise QTextDocument (HTML) + QPrinter en mode sortie fichier.
        """
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter PDF",
            _re.sub(r'[\\/:*?"<>|]', '_', f"chiffrage_{self.matiere_courante}_{datetime.date.today()}.pdf"),
            "PDF (*.pdf)"
        )
        if not path:
            return
        clients, all_years = self._rows_export()
        mat     = self.matiere_courante
        pl21_txt = f" | PL21 : {self._pl21:.2f} €" if self._pl21 else ""

        # En-têtes du tableau
        th = "".join(
            f"<th style='padding:5px;'>{h}</th>"
            for h in (
                ["Client"] +
                [f"Taux {y}" for y in all_years] +
                ["Remise"] +
                (["Prix av. remise", "Prix final"] if self._pl21 else [])
            )
        )
        # Lignes du tableau (palette fixe, indépendante du thème écran). Les noms
        # de client sont du texte libre saisi par un admin : échappés (html.escape)
        # pour ne pas casser le rendu HTML si l'un d'eux contient &, < ou >.
        pp = print_palette(mat)
        rows_html = ""
        for idx, client in enumerate(clients):
            taux_c = DATA["augmentations"][mat][client]
            remise = DATA["remises"][mat][client]
            bg     = pp["stripe"] if idx % 2 else "#ffffff"
            tds    = f"<td style='padding:4px;'>{_html.escape(client)}</td>"
            for y in all_years:
                t = taux_c.get(y, 0)
                tds += f"<td align='center'>+{t * 100:.1f}%</td>"
            tds += f"<td align='center' style='color:{pp['danger_fg']};'>{_fmt_remise(abs(remise) * 100)}</td>"
            if self._pl21:
                res  = calculer_prix(self._pl21, mat, client)
                tds += (
                    f"<td align='center'>{res['prix_avant_remise']:.2f} €</td>"
                    f"<td align='center' style='color:{pp['success_fg']};font-weight:bold;'>"
                    f"{res['prix_final']} €</td>"
                )
            rows_html += f"<tr style='background:{bg};'>{tds}</tr>"

        html = (
            f"<html><body style='font-family:Arial;font-size:9pt;color:{pp['ink']};'>"
            f"<h2 style='color:{pp['accent']};'>Chiffrage OEM — {_html.escape(mat)}</h2>"
            # Récapitulatif agrégeant les remises de tous les clients (contrairement
            # à l'export PDF « client », propre à un seul destinataire).
            f"<p style='color:#808080;font-style:italic;'>USAGE INTERNE — NE PAS DIFFUSER</p>"
            f"<p>Date : {datetime.date.today().strftime('%d/%m/%Y')}{pl21_txt}</p>"
            f"<table border='1' cellspacing='0' style='border-collapse:collapse;width:100%;border-color:{pp['line']};'>"
            f"<tr style='background:{pp['accent']};color:white;'>{th}</tr>"
            f"{rows_html}</table></body></html>"
        )
        if export_html_to_pdf(self, html, path):
            QMessageBox.information(self, "Export PDF", f"Exporté :\n{path}")

    # ── Actions administrateur ────────────────────────────────────────────────

    def _admin_ajouter_client(self):
        """Ouvre AddClientDialog (après vérification admin) et recharge la sidebar."""
        if not verifier_admin(self):
            return
        dlg = AddClientDialog(self.matiere_courante, self)
        accepted = dlg.exec() == QDialog.DialogCode.Accepted
        nom = dlg.get_nom()
        dlg.deleteLater()
        if accepted:
            self._populate_clients(self.matiere_courante)
            QMessageBox.information(self, "Client ajouté", f"'{nom}' ajouté.")

    def _admin_supprimer_client(self):
        """
        Supprime un client de TOUTES les matières après confirmation.
        La suppression est irréversible (sauf restauration manuelle du JSON).
        """
        if not verifier_admin(self):
            return
        clients = list(DATA["augmentations"][self.matiere_courante].keys())
        if not clients:
            return
        client, ok = QInputDialog.getItem(
            self, "Supprimer un client", "Client :", clients, 0, False
        )
        if not ok:
            return
        if QMessageBox.question(
            self, "Confirmer",
            f"Supprimer '{client}' de toutes les matières ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) != QMessageBox.StandardButton.Yes:
            return
        for mat in DATA["augmentations"]:
            DATA["augmentations"][mat].pop(client, None)
            DATA["remises"].get(mat, {}).pop(client, None)
        DATA.get("remarques", {}).pop(client, None)
        # log_audit() persiste DATA : pas besoin d'un save_data() séparé ici.
        log_audit("delete_client", self.matiere_courante, client)
        self._populate_clients(self.matiere_courante)
        QMessageBox.information(self, "Supprimé", f"'{client}' supprimé.")

    def _admin_ajouter_matiere(self):
        """
        Crée une nouvelle matière avec les mêmes clients que la première matière existante
        (tous les taux initialisés à 0 %, remises à -60 %).
        """
        if not verifier_admin(self):
            return
        nom, ok = QInputDialog.getText(self, "Nouvelle matière", "Nom de la matière :")
        if not ok or not nom.strip():
            return
        nom = nom.strip().upper()
        if nom in DATA["augmentations"]:
            QMessageBox.warning(self, "Erreur", f"'{nom}' existe déjà.")
            return
        # DATA["augmentations"] est normalement toujours non vide (PTFE/PEEK par
        # défaut, aucune fonction ne supprime une matière) ; garde défensive au
        # cas où un fichier de données manipulé à la main l'aurait vidée.
        ref = next(iter(DATA["augmentations"]), None)
        if ref is None:
            DATA["augmentations"][nom] = {}
            DATA["remises"][nom] = {}
        else:
            # 'augmentations[ref]' est l'unique source de vérité pour la liste
            # des clients : 'remises[ref]' est lu avec un repli (-60%) plutôt
            # que d'itérer ses propres clés, pour ne pas propager/amplifier une
            # éventuelle désynchronisation entre les deux dicts (ex: fichier
            # édité à la main où un client manquerait côté remises).
            clients_ref = DATA["augmentations"][ref]
            remises_ref = DATA["remises"].get(ref, {})
            DATA["augmentations"][nom] = {
                c: {a: 0.0 for a in t}
                for c, t in clients_ref.items()
            }
            DATA["remises"][nom] = {
                c: remises_ref[c] if isinstance(remises_ref.get(c), (int, float)) else -0.60
                for c in clients_ref
            }
        # log_audit() persiste DATA : pas besoin d'un save_data() séparé ici.
        log_audit("add_matiere", nom)
        self._create_matiere_btn(nom)
        QMessageBox.information(self, "Matière ajoutée", f"'{nom}' ajoutée.")

    def _admin_changer_mdp(self):
        """Ouvre ChangePasswordDialog (vérification de l'ancien mot de passe d'abord)."""
        if not verifier_admin(self):
            return
        dlg = ChangePasswordDialog(self)
        dlg.exec()
        dlg.deleteLater()


# ─────────────────────────────────────────────────────────────────────────────
#  POINT D'ENTRÉE
# ─────────────────────────────────────────────────────────────────────────────

def _excepthook(exc_type, exc_value, exc_tb):
    """
    Filet de sécurité global : PyQt6 abandonne (abort) tout le processus si une
    exception Python traverse un slot connecté à un signal Qt (ex: clic sur
    « Calculer »), ce qui se manifeste pour l'utilisateur comme un plantage sec
    de l'application sans message. On affiche donc l'erreur dans une boîte de
    dialogue et on laisse l'application continuer à tourner.

    Le build packagé tourne en mode --windowed (pas de console) : print(...,
    file=sys.stderr) n'est alors visible nulle part. La trace complète est
    donc aussi ajoutée à un fichier journal à côté des données, pour qu'un
    incident en production reste diagnosticable après coup.
    """
    import traceback
    msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print(msg, file=sys.stderr)
    try:
        log_path = _data_dir() / "chiffrage_crash.log"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.datetime.now().isoformat(timespec='seconds')}]\n{msg}")
    except OSError:
        pass
    try:
        QMessageBox.critical(
            None, "Erreur inattendue",
            "Une erreur inattendue s'est produite et a été ignorée pour éviter "
            f"un plantage :\n\n{exc_value}"
        )
    except Exception:
        pass


def main():
    """
    Point d'entrée de l'application.
    - Crée la QApplication avec le style Fusion (rendu cohérent sur Windows/Mac/Linux)
    - Crée et affiche la fenêtre principale : c'est MainWindow._apply_theme() qui
      définit la palette complète (thème clair par défaut, bascule possible en sombre
      via le bouton lune/soleil du header)
    """
    sys.excepthook = _excepthook
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Barre de titre Windows thémée pour toutes les fenêtres (dialogues compris)
    titlebar_themer = TitlebarThemer(app)
    app.installEventFilter(titlebar_themer)

    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
